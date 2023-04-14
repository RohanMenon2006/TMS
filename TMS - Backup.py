"""The Time Management System - C.S Project"""

# Importing required modules
import customtkinter as ctk
import tkinter
from tkinter import ttk
from PIL import ImageTk
from PIL import Image
import openpyxl

# importing the setting of the program
fam = open("appearance.txt", "r")
appearance = fam.read()
fam.close()

fsm = open('Scale.txt', 'r')
scaling = float(fsm.read())
fsm.close()
file_scaling = int(scaling) / 100

# setting the theme of the program
ctk.set_appearance_mode(appearance)
ctk.set_default_color_theme("green")
ctk.set_widget_scaling(file_scaling)

# making the program fullscreen only
root = ctk.CTk()
root.overrideredirect(True)
root.geometry("{0}x{1}+0+0".format(root.winfo_screenwidth(), root.winfo_screenheight()))

# setting the frame of the program
frame = ctk.CTkFrame(master=root, corner_radius=15)
frame.pack(fill="both", expand=True)

# dev team name x_x
name = "The Mallus"

# defined y location for the login form to make adjustment easier
login_form_y = 0.45


# login program
# if rl is false is runs the login program
def tab1():
    label = ctk.CTkLabel(master=frame, text="Login Page", font=('Segoe Ui', 45))
    label.pack(fill="both", expand=True)
    label.place(relx=0.5, rely=0.08, anchor=ctk.N)

    frame1 = ctk.CTkFrame(master=frame, width=320, height=360, corner_radius=20)
    frame1.place(relx=0.5, rely=0.5, anchor=tkinter.CENTER)

    label1 = ctk.CTkLabel(master=frame1, text="Log into your Account", font=('Segoe Ui', 25))
    label1.pack()
    label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

    entry1 = ctk.CTkEntry(master=frame1, width=250, placeholder_text="Username", font=('Segoe Ui', 17))
    entry1.pack(padx=10, pady=12)
    entry1.place(relx=0.5, rely=login_form_y - 0.08, anchor=ctk.CENTER)

    entry2 = ctk.CTkEntry(master=frame1, width=250, placeholder_text="Password", show="*", font=('Segoe Ui', 17))
    entry2.pack(padx=10, pady=12)
    entry2.place(relx=0.5, rely=login_form_y + 0.07, anchor=ctk.CENTER)

    cb_va = ctk.IntVar(value=0)

    def my_show():
        if cb_va.get() == 1:
            entry2.configure(show='')
        else:
            entry2.configure(show='*')

    checkbox1 = ctk.CTkCheckBox(master=frame1, text='', corner_radius=5, font=('Segoe Ui', 12), variable=cb_va,
                                onvalue=1, offvalue=0, command=my_show)
    checkbox1.configure(width=0, height=1)
    checkbox1.place(relx=0.836, rely=login_form_y + 0.07, anchor=ctk.CENTER)

    def p_forgot():
        textbox4
        textbox4.place(relx=0.5, rely=login_form_y + 0.3, anchor=ctk.CENTER)
        textbox3.place(relx=2, rely=2)

    forgot = ctk.CTkButton(master=frame1, text="Forgot Password?", font=('Century Gothic', 12), fg_color="#333333",
                           command=p_forgot)
    forgot.configure(height=1, width=1)
    forgot.place(relx=0.75, rely=login_form_y + 0.2, anchor=ctk.CENTER)

    check_state = ctk.IntVar()

    def check_login():
        if entry1.get() == "Admin" and entry2.get() == "Admin123":
            if check_state.get() == 1:
                fh = open('login_remember_admin.txt', 'w')
                str1 = "true"
                print(str1)
                fh.write(str1)
                fh.close()
            elif check_state.get() == 0:
                fh = open('login_remember_admin.txt', 'w')
                str2 = "false"
                print(str2)
                fh.write(str2)
                fh.close()
        if entry1.get() == "Teacher" and entry2.get() == "Teacher123":
            if check_state.get() == 1:
                fh = open('login remember teach.txt', 'w')
                str1 = "true"
                print(str1)
                fh.write(str1)
                fh.close()
            elif check_state.get() == 0:
                fh = open('login remember teach.txt', 'w')
                str2 = "false"
                print(str2)
                fh.write(str2)
                fh.close()

    checkbox2 = ctk.CTkCheckBox(master=frame1, text="Remember Me", font=('Century Gothic', 12), variable=check_state,
                                command=check_login)
    checkbox2.pack(padx=10, pady=12)
    checkbox2.configure(height=0, width=0)
    checkbox2.place(relx=0.3, rely=login_form_y + 0.2, anchor=ctk.CENTER)

    textbox3 = ctk.CTkLabel(master=frame1, height=5, text="*Incorrect Username or Password*", font=('Segoe Ui', 13))

    textbox4 = ctk.CTkLabel(master=frame1, height=5, text="How did you forget that? :|", font=('Segoe Ui', 13))

    def destory_login():
        label.destroy()
        label1.destroy()
        frame1.configure(width=0, height=0)
        entry1.destroy()
        entry2.destroy()
        checkbox1.destroy()
        checkbox2.destroy()
        textbox3.destroy()
        textbox4.destroy()
        forgot.destroy()
        cbutton.destroy()
        _quit.destroy()
        text_name.destroy()
        tab2()

    def destory_login2():
        label.destroy()
        label1.destroy()
        frame1.configure(width=0, height=0)
        entry1.destroy()
        entry2.destroy()
        checkbox1.destroy()
        checkbox2.destroy()
        textbox3.destroy()
        textbox4.destroy()
        forgot.destroy()
        cbutton.destroy()
        _quit.destroy()
        text_name.destroy()
        ttab2()

    def login_details():
        if entry1.get() == "Admin" and entry2.get() == "Admin123":
            destory_login()
        elif entry1.get() == "Teacher" and entry2.get() == "Teacher123":
            destory_login2()
        else:
            textbox3
            textbox3.place(relx=0.5, rely=login_form_y + 0.3, anchor=ctk.CENTER)
            textbox4.place(relx=2, rely=2)

    cbutton = ctk.CTkButton(master=frame1, width=180, height=15, corner_radius=100, text="Login", font=('Segoe Ui', 20),
                            command=login_details)
    cbutton.pack(padx=10, pady=12)
    cbutton.place(relx=0.5, rely=login_form_y + 0.43, anchor=ctk.CENTER)

    _quit = ctk.CTkButton(master=frame, text="Quit", font=('Segoe Ui', 15), command=exit)
    _quit.pack()
    _quit.place(relx=0.89, rely=0.95, anchor=ctk.NW)

    text_name = ctk.CTkLabel(master=frame, height=5, text=name, corner_radius=15, font=('Segoe Ui', 18))
    text_name.pack()
    text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)


# program lobby - Admin
# if rl is true it skips login and runs the program
def tab2():
    label1 = ctk.CTkLabel(master=frame, text="Time Management System", font=('Segoe Ui', 55))
    label1.pack(fill="both", expand=True)
    label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

    acb_va = ctk.IntVar(value=0)

    # noinspection PyUnboundLocalVariable,PyGlobalUndefined
    def account():
        if acb_va.get() == 1:

            global admin_frame

            admin_frame = ctk.CTkFrame(master=frame, width=150, height=150, corner_radius=15)
            admin_frame.pack()
            admin_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

            def settings():
                settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                settings_frame.pack(fill="both", expand=True)

                settings_1 = ctk.CTkFrame(master=settings_frame, corner_radius=30)
                settings_1.pack(pady=60, padx=550, fill="both", expand=True)

                settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                settings_label.pack(padx=10, pady=12)
                settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                def change_appearance_mode_event(new_appearance_mode: str):
                    ctk.set_appearance_mode(new_appearance_mode)
                    fa = open("appearance.txt", "w")
                    fa.write(new_appearance_mode)
                    fa.close()

                appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:", font=('Segoe Ui', 20))
                appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                           command=change_appearance_mode_event, font=('Segoe Ui', 15))
                appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                def change_scaling_event(new_scaling: str):
                    new_scaling_float = int(new_scaling.replace("%", "")) / 100
                    ctk.set_widget_scaling(new_scaling_float)
                    fs = open('Scale.txt', 'w')
                    n_scaling = new_scaling.replace("%", "")
                    fs.write(n_scaling)
                    fs.close()

                scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:", font=('Segoe Ui', 20))
                scaling_label.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)
                scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                   command=change_scaling_event, font=('Segoe Ui', 15))
                scaling_option.place(relx=0.5, rely=0.6, anchor=ctk.CENTER)

                def destroy_settings():
                    settings_frame.destroy()
                    settings_1.destroy()
                    settings_label.destroy()
                    appearance_mode_label.destroy()
                    appearance_mode_option.destroy()
                    scaling_label.destroy()
                    scaling_option.destroy()
                    _save.destroy()

                _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15), command=destroy_settings)
                _save.pack(padx=10, pady=12)
                _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

            def logout():
                fl = open('login_remember_admin.txt', 'w')
                str2 = "false"
                print(str2)
                fl.write(str2)
                fl.close()
                destroy_account()
                destroy_tab2a()
                tab1()

            def destroy_account():
                admin_frame.destroy()
                settings_button.destroy()
                logout1.destroy()
                _quit.destroy()

            def destroy_tab2a():
                label1.destroy()
                view_button.destroy()
                man_button.destroy()
                sub_button.destroy()
                admin_label.destroy()
                admin_checkbox.destroy()
                logout1.destroy()
                text_name.destroy()
                _quit.destroy()

            settings_button = ctk.CTkButton(master=admin_frame, text="Settings", font=('Segoe Ui', 15),
                                            command=settings)
            settings_button.pack(padx=10, pady=12)
            settings_button.place(relx=0.5, rely=0.25, anchor=ctk.CENTER)

            logout1 = ctk.CTkButton(master=admin_frame, text="Logout", font=('Segoe Ui', 15), command=logout)
            logout1.pack(padx=10, pady=10)
            logout1.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            _quit = ctk.CTkButton(master=admin_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
            _quit.pack(padx=10, pady=12)
            _quit.place(relx=0.5, rely=0.75, anchor=ctk.CENTER)
        else:
            admin_frame.destroy()

    admin_checkbox = ctk.CTkCheckBox(master=frame, text='Admin', corner_radius=5, font=('Segoe Ui', 21),
                                     variable=acb_va, onvalue=1, offvalue=0, command=account)
    admin_checkbox.pack()
    admin_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

    user_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

    admin_label = ctk.CTkLabel(master=frame, text="", image=user_image)
    admin_label.pack()
    admin_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

    def destroy_tab2():
        label1.destroy()
        view_button.destroy()
        man_button.destroy()
        sub_button.destroy()
        admin_label.destroy()
        admin_checkbox.destroy()
        text_name.destroy()
        if acb_va.get() == 1:
            admin_frame.destroy()

    def button_view():
        destroy_tab2()
        tabview()

    def button_man():
        destroy_tab2()
        tabman()

    def button_sub():
        destroy_tab2()
        tabsub()

    view_button = ctk.CTkButton(frame, text="View Timetables", font=('Segoe Ui', 45), command=button_view)
    view_button.pack(padx=10, pady=30)
    view_button.place(relx=0.5, rely=0.35, anchor=ctk.N)

    man_button = ctk.CTkButton(frame, text="Manage Timetables", font=('Segoe Ui', 45), command=button_man)
    man_button.pack(padx=10, pady=30)
    man_button.place(relx=0.5, rely=0.5, anchor=ctk.N)

    sub_button = ctk.CTkButton(frame, text="Substitute Absent Teachers", font=('Segoe Ui', 45), command=button_sub)
    sub_button.pack(padx=10, pady=30)
    sub_button.place(relx=0.5, rely=0.65, anchor=ctk.N)

    text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
    text_name.pack(padx=1, pady=2)
    text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)


# to view the timetable - Admin
def tabview():
    view_label = ctk.CTkLabel(master=frame, text="Timetables", font=('Segoe Ui', 40))
    view_label.pack()
    view_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

    view_label2 = ctk.CTkLabel(master=frame, text="Pick a class", font=('Segoe Ui', 35))
    view_label2.pack()
    view_label2.place(relx=0.5, rely=0.2, anchor=ctk.N)

    acb_va = ctk.IntVar(value=0)

    # noinspection PyUnboundLocalVariable,PyGlobalUndefined
    def account():
        if acb_va.get() == 1:

            global admin_frame

            admin_frame = ctk.CTkFrame(master=frame, width=150, height=100, corner_radius=15)
            admin_frame.pack()
            admin_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

            def settings():
                settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                settings_frame.pack(fill="both", expand=True)

                settings_1 = ctk.CTkFrame(master=settings_frame, corner_radius=30)
                settings_1.pack(pady=60, padx=550, fill="both", expand=True)

                settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                settings_label.pack(padx=10, pady=12)
                settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                def change_appearance_mode_event(new_appearance_mode: str):
                    ctk.set_appearance_mode(new_appearance_mode)
                    fa = open("appearance.txt", "w")
                    fa.write(new_appearance_mode)
                    fa.close()

                appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:", font=('Segoe Ui', 20))
                appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                           command=change_appearance_mode_event, font=('Segoe Ui', 15))
                appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                def change_scaling_event(new_scaling: str):
                    new_scaling_float = int(new_scaling.replace("%", "")) / 100
                    ctk.set_widget_scaling(new_scaling_float)
                    fs = open('Scale.txt', 'w')
                    n_scaling = new_scaling.replace("%", "")
                    fs.write(n_scaling)
                    fs.close()

                scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:", font=('Segoe Ui', 20))
                scaling_label.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)
                scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                   command=change_scaling_event, font=('Segoe Ui', 15))
                scaling_option.place(relx=0.5, rely=0.6, anchor=ctk.CENTER)

                def destroy_settings():
                    settings_frame.destroy()
                    settings_1.destroy()
                    settings_label.destroy()
                    appearance_mode_label.destroy()
                    appearance_mode_option.destroy()
                    scaling_label.destroy()
                    scaling_option.destroy()
                    _save.destroy()

                _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15), command=destroy_settings)
                _save.pack(padx=10, pady=12)
                _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

            settings_button = ctk.CTkButton(master=admin_frame, text="Settings", font=('Segoe Ui', 15),
                                            command=settings)
            settings_button.pack(padx=10, pady=12)
            settings_button.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

            _quit = ctk.CTkButton(master=admin_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
            _quit.pack(padx=10, pady=12)
            _quit.place(relx=0.5, rely=0.7, anchor=ctk.CENTER)
        else:
            admin_frame.destroy()

    admin_checkbox = ctk.CTkCheckBox(master=frame, text='Admin', corner_radius=5, font=('Segoe Ui', 21),
                                     variable=acb_va, onvalue=1, offvalue=0, command=account)
    admin_checkbox.pack()
    admin_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

    user_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

    admin_label = ctk.CTkLabel(master=frame, text="", image=user_image)
    admin_label.pack()
    admin_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

    # reading the Excel files

    def class11():

        class11_label = ctk.CTkLabel(master=frame, text="Timetables", font=('Segoe Ui', 40))
        class11_label.pack()
        class11_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

        class11_label2 = ctk.CTkLabel(master=frame, text="Class 11", font=('Segoe Ui', 35))
        class11_label2.pack()
        class11_label2.place(relx=0.5, rely=0.2, anchor=ctk.N)

        acb_va1 = ctk.IntVar(value=0)

        # noinspection PyUnboundLocalVariable,PyGlobalUndefined
        def account1():
            if acb_va1.get() == 1:

                global admin_frame

                admin_frame = ctk.CTkFrame(master=frame, width=150, height=100, corner_radius=15)
                admin_frame.pack()
                admin_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

                def settings():
                    settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                    settings_frame.pack(fill="both", expand=True)

                    settings_1 = ctk.CTkFrame(master=settings_frame)
                    settings_1.pack(pady=20, padx=60, fill="both", expand=True)

                    settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                    settings_label.pack(padx=10, pady=12)
                    settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                    def change_appearance_mode_event(new_appearance_mode: str):
                        ctk.set_appearance_mode(new_appearance_mode)
                        fa = open("appearance.txt", "w")
                        fa.write(new_appearance_mode)
                        fa.close()

                    appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:")
                    appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                    appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                               command=change_appearance_mode_event)
                    appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                    def change_scaling_event(new_scaling: str):
                        new_scaling_float = int(new_scaling.replace("%", "")) / 100
                        ctk.set_widget_scaling(new_scaling_float)
                        fs = open('Scale.txt', 'w')
                        n_scaling = new_scaling.replace("%", "")
                        fs.write(n_scaling)
                        fs.close()

                    scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:")
                    scaling_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)
                    scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                       command=change_scaling_event)
                    scaling_option.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)

                    def destroy_settings():
                        settings_frame.destroy()
                        settings_1.destroy()
                        settings_label.destroy()
                        appearance_mode_label.destroy()
                        appearance_mode_option.destroy()
                        scaling_label.destroy()
                        scaling_option.destroy()
                        _save.destroy()

                    _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15),
                                          command=destroy_settings)
                    _save.pack(padx=10, pady=12)
                    _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

                settings_button = ctk.CTkButton(master=admin_frame, text="Settings", font=('Segoe Ui', 15),
                                                command=settings)
                settings_button.pack(padx=10, pady=12)
                settings_button.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

                _quit = ctk.CTkButton(master=admin_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
                _quit.pack(padx=10, pady=12)
                _quit.place(relx=0.5, rely=0.7, anchor=ctk.CENTER)
            else:
                admin_frame.destroy()

        admin1_checkbox = ctk.CTkCheckBox(master=frame, text='Admin', corner_radius=5, font=('Segoe Ui', 21),
                                          variable=acb_va1, onvalue=1, offvalue=0, command=account1)
        admin1_checkbox.pack()
        admin1_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

        user1_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

        admin1_label = ctk.CTkLabel(master=frame, text="", image=user1_image)
        admin1_label.pack()
        admin1_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

        def view1():

            view1_label = ctk.CTkLabel(master=frame, text="Timetables", font=('Segoe Ui', 40))
            view1_label.pack()
            view1_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

            class_label = ctk.CTkLabel(master=frame, text="Class 11 A", font=('Segoe Ui', 35))
            class_label.pack()
            class_label.place(relx=0.5, rely=0.2, anchor=ctk.N)

            acb_va2 = ctk.IntVar(value=0)

            # noinspection PyUnboundLocalVariable,PyGlobalUndefined
            def account2():
                if acb_va2.get() == 1:

                    global admin_frame

                    admin_frame = ctk.CTkFrame(master=frame, width=150, height=100, corner_radius=15)
                    admin_frame.pack()
                    admin_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

                    def settings():
                        settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                        settings_frame.pack(fill="both", expand=True)

                        settings_1 = ctk.CTkFrame(master=settings_frame)
                        settings_1.pack(pady=20, padx=60, fill="both", expand=True)

                        settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                        settings_label.pack(padx=10, pady=12)
                        settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                        def change_appearance_mode_event(new_appearance_mode: str):
                            ctk.set_appearance_mode(new_appearance_mode)
                            fa = open("appearance.txt", "w")
                            fa.write(new_appearance_mode)
                            fa.close()

                        appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:")
                        appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                        appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                                   command=change_appearance_mode_event)
                        appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                        def change_scaling_event(new_scaling: str):
                            new_scaling_float = int(new_scaling.replace("%", "")) / 100
                            ctk.set_widget_scaling(new_scaling_float)
                            fs = open('Scale.txt', 'w')
                            n_scaling = new_scaling.replace("%", "")
                            fs.write(n_scaling)
                            fs.close()

                        scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:")
                        scaling_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)
                        scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                           command=change_scaling_event)
                        scaling_option.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)

                        def destroy_settings():
                            settings_frame.destroy()
                            settings_1.destroy()
                            settings_label.destroy()
                            appearance_mode_label.destroy()
                            appearance_mode_option.destroy()
                            scaling_label.destroy()
                            scaling_option.destroy()
                            _save.destroy()

                        _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15),
                                              command=destroy_settings)
                        _save.pack(padx=10, pady=12)
                        _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

                    settings_button = ctk.CTkButton(master=admin_frame, text="Settings", font=('Segoe Ui', 15),
                                                    command=settings)
                    settings_button.pack(padx=10, pady=12)
                    settings_button.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

                    _quit = ctk.CTkButton(master=admin_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
                    _quit.pack(padx=10, pady=12)
                    _quit.place(relx=0.5, rely=0.7, anchor=ctk.CENTER)
                else:
                    admin_frame.destroy()

            admin2_checkbox = ctk.CTkCheckBox(master=frame, text='Admin', corner_radius=5, font=('Segoe Ui', 21),
                                              variable=acb_va2, onvalue=1, offvalue=0, command=account2)
            admin2_checkbox.pack()
            admin2_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

            user2_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

            admin2_label = ctk.CTkLabel(master=frame, text="", image=user2_image)
            admin2_label.pack()
            admin2_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

            path = "timetable_teachers.xlsx"
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook['Class 11 A']
            header = sheet.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            body = sheet.iter_rows(min_row=2, max_row=6, max_col=9, values_only=True)
            header = [r for r in header]
            body = [r for r in body]
            workbook.close()

            style = ttk.Style()
            style.configure("Treeview", rowheight=60)
            style.configure("Treeview.Heading", font=('Segoe UI', 20))

            tree = ttk.Treeview(master=frame, selectmode='browse', show='headings', columns=header[0], height=5)
            tree.pack(padx=10, pady=10, fill='x')
            tree.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            tree.tag_configure("tag_item", font=('Segoe Ui', 12), foreground='black')

            for i in header[0]:
                tree.column(i, anchor=ctk.CENTER)
                tree.heading(i, text=i)

            for i in body:
                tree.insert('', 'end', iid=i[0], values=i, tags="tag_item")

            text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
            text_name.pack(padx=1, pady=2)
            text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

            def _back():
                view1_label.destroy()
                view1_back.destroy()
                text_name.destroy()
                class_label.destroy()
                tree.destroy()
                admin2_checkbox.destroy()
                admin2_label.destroy()
                if acb_va2.get() == 1:
                    admin_frame.destroy()
                class11()

            view1_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
            view1_back.pack(pady=60)
            view1_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

        def view2():

            view2_label = ctk.CTkLabel(master=frame, text="Timetables", font=('Segoe Ui', 40))
            view2_label.pack()
            view2_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

            class_label = ctk.CTkLabel(master=frame, text="Class 11 C", font=('Segoe Ui', 35))
            class_label.pack()
            class_label.place(relx=0.5, rely=0.2, anchor=ctk.N)

            acb_va2 = ctk.IntVar(value=0)

            # noinspection PyUnboundLocalVariable,PyGlobalUndefined
            def account2():
                if acb_va2.get() == 1:

                    global admin_frame

                    admin_frame = ctk.CTkFrame(master=frame, width=150, height=100, corner_radius=15)
                    admin_frame.pack()
                    admin_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

                    def settings():
                        settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                        settings_frame.pack(fill="both", expand=True)

                        settings_1 = ctk.CTkFrame(master=settings_frame)
                        settings_1.pack(pady=20, padx=60, fill="both", expand=True)

                        settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                        settings_label.pack(padx=10, pady=12)
                        settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                        def change_appearance_mode_event(new_appearance_mode: str):
                            ctk.set_appearance_mode(new_appearance_mode)
                            fa = open("appearance.txt", "w")
                            fa.write(new_appearance_mode)
                            fa.close()

                        appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:")
                        appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                        appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                                   command=change_appearance_mode_event)
                        appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                        def change_scaling_event(new_scaling: str):
                            new_scaling_float = int(new_scaling.replace("%", "")) / 100
                            ctk.set_widget_scaling(new_scaling_float)
                            fs = open('Scale.txt', 'w')
                            n_scaling = new_scaling.replace("%", "")
                            fs.write(n_scaling)
                            fs.close()

                        scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:")
                        scaling_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)
                        scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                           command=change_scaling_event)
                        scaling_option.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)

                        def destroy_settings():
                            settings_frame.destroy()
                            settings_1.destroy()
                            settings_label.destroy()
                            appearance_mode_label.destroy()
                            appearance_mode_option.destroy()
                            scaling_label.destroy()
                            scaling_option.destroy()
                            _save.destroy()

                        _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15),
                                              command=destroy_settings)
                        _save.pack(padx=10, pady=12)
                        _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

                    settings_button = ctk.CTkButton(master=admin_frame, text="Settings", font=('Segoe Ui', 15),
                                                    command=settings)
                    settings_button.pack(padx=10, pady=12)
                    settings_button.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

                    _quit = ctk.CTkButton(master=admin_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
                    _quit.pack(padx=10, pady=12)
                    _quit.place(relx=0.5, rely=0.7, anchor=ctk.CENTER)
                else:
                    admin_frame.destroy()

            admin2_checkbox = ctk.CTkCheckBox(master=frame, text='Admin', corner_radius=5, font=('Segoe Ui', 21),
                                              variable=acb_va2, onvalue=1, offvalue=0, command=account2)
            admin2_checkbox.pack()
            admin2_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

            user2_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

            admin2_label = ctk.CTkLabel(master=frame, text="", image=user2_image)
            admin2_label.pack()
            admin2_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

            path = "timetable_teachers.xlsx"
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook['Class 11 B']
            header = sheet.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            body = sheet.iter_rows(min_row=2, max_row=6, max_col=9, values_only=True)
            header = [r for r in header]
            body = [r for r in body]
            workbook.close()

            style = ttk.Style()
            style.configure("Treeview", rowheight=60)
            style.configure("Treeview.Heading", font=('Segoe UI', 20))

            tree = ttk.Treeview(master=frame, selectmode='browse', show='headings', columns=header[0], height=5)
            tree.pack(padx=10, pady=10, fill='x')
            tree.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            tree.tag_configure("tag_item", font=('Segoe Ui', 12), foreground='black')

            for i in header[0]:
                tree.column(i, anchor=ctk.CENTER)
                tree.heading(i, text=i)

            for i in body:
                tree.insert('', 'end', iid=i[0], values=i, tags="tag_item")

            text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
            text_name.pack(padx=1, pady=2)
            text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

            def _back():
                view2_label.destroy()
                view2_back.destroy()
                text_name.destroy()
                class_label.destroy()
                tree.destroy()
                admin2_checkbox.destroy()
                admin2_label.destroy()
                if acb_va2.get() == 1:
                    admin_frame.destroy()
                class11()

            view2_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
            view2_back.pack(pady=60)
            view2_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

        def go_view1():
            class11_label.destroy()
            class11_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class11_name.destroy()
            class11_back.destroy()
            admin1_checkbox.destroy()
            admin1_label.destroy()
            if acb_va1.get() == 1:
                admin_frame.destroy()
            view1()

        def go_view2():
            class11_label.destroy()
            class11_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class11_name.destroy()
            class11_back.destroy()
            admin1_label.destroy()
            admin1_checkbox.destroy()
            if acb_va1.get() == 1:
                admin_frame.destroy()
            view2()

        view1_button = ctk.CTkButton(frame, text="Class 11 A", font=('Segoe Ui', 25), command=go_view1)
        view1_button.pack(padx=10, pady=30)
        view1_button.place(relx=0.5, rely=0.45, anchor=ctk.N)

        view2_button = ctk.CTkButton(frame, text="Class 11 C", font=('Segoe Ui', 25), command=go_view2)
        view2_button.pack(padx=10, pady=30)
        view2_button.place(relx=0.5, rely=0.55, anchor=ctk.N)

        def back11():
            class11_label.destroy()
            class11_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class11_name.destroy()
            class11_back.destroy()
            admin1_label.destroy()
            admin1_checkbox.destroy()
            if acb_va1.get() == 1:
                admin_frame.destroy()
            tabview()

        class11_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
        class11_name.pack(padx=1, pady=2)
        class11_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

        class11_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=back11)
        class11_back.pack()
        class11_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

    def class10():

        class10_label = ctk.CTkLabel(master=frame, text="Timetables", font=('Segoe Ui', 40))
        class10_label.pack()
        class10_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

        class10_label2 = ctk.CTkLabel(master=frame, text="Class 10", font=('Segoe Ui', 35))
        class10_label2.pack()
        class10_label2.place(relx=0.5, rely=0.2, anchor=ctk.N)

        acb_va1 = ctk.IntVar(value=0)

        # noinspection PyUnboundLocalVariable,PyGlobalUndefined
        def account1():
            if acb_va1.get() == 1:

                global admin_frame

                admin_frame = ctk.CTkFrame(master=frame, width=150, height=100, corner_radius=15)
                admin_frame.pack()
                admin_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

                def settings():
                    settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                    settings_frame.pack(fill="both", expand=True)

                    settings_1 = ctk.CTkFrame(master=settings_frame)
                    settings_1.pack(pady=20, padx=60, fill="both", expand=True)

                    settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                    settings_label.pack(padx=10, pady=12)
                    settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                    def change_appearance_mode_event(new_appearance_mode: str):
                        ctk.set_appearance_mode(new_appearance_mode)
                        fa = open("appearance.txt", "w")
                        fa.write(new_appearance_mode)
                        fa.close()

                    appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:")
                    appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                    appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                               command=change_appearance_mode_event)
                    appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                    def change_scaling_event(new_scaling: str):
                        new_scaling_float = int(new_scaling.replace("%", "")) / 100
                        ctk.set_widget_scaling(new_scaling_float)
                        fs = open('Scale.txt', 'w')
                        n_scaling = new_scaling.replace("%", "")
                        fs.write(n_scaling)
                        fs.close()

                    scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:")
                    scaling_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)
                    scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                       command=change_scaling_event)
                    scaling_option.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)

                    def destroy_settings():
                        settings_frame.destroy()
                        settings_1.destroy()
                        settings_label.destroy()
                        appearance_mode_label.destroy()
                        appearance_mode_option.destroy()
                        scaling_label.destroy()
                        scaling_option.destroy()
                        _save.destroy()

                    _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15),
                                          command=destroy_settings)
                    _save.pack(padx=10, pady=12)
                    _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

                settings_button = ctk.CTkButton(master=admin_frame, text="Settings", font=('Segoe Ui', 15),
                                                command=settings)
                settings_button.pack(padx=10, pady=12)
                settings_button.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

                _quit = ctk.CTkButton(master=admin_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
                _quit.pack(padx=10, pady=12)
                _quit.place(relx=0.5, rely=0.7, anchor=ctk.CENTER)
            else:
                admin_frame.destroy()

        admin1_checkbox = ctk.CTkCheckBox(master=frame, text='Admin', corner_radius=5, font=('Segoe Ui', 21),
                                          variable=acb_va1, onvalue=1, offvalue=0, command=account1)
        admin1_checkbox.pack()
        admin1_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

        user1_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

        admin1_label = ctk.CTkLabel(master=frame, text="", image=user1_image)
        admin1_label.pack()
        admin1_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

        def view1():

            view1_label = ctk.CTkLabel(master=frame, text="Timetables", font=('Segoe Ui', 40))
            view1_label.pack()
            view1_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

            class_label = ctk.CTkLabel(master=frame, text="Class 11 A", font=('Segoe Ui', 35))
            class_label.pack()
            class_label.place(relx=0.5, rely=0.2, anchor=ctk.N)

            acb_va2 = ctk.IntVar(value=0)

            # noinspection PyUnboundLocalVariable,PyGlobalUndefined
            def account2():
                if acb_va2.get() == 1:

                    global admin_frame

                    admin_frame = ctk.CTkFrame(master=frame, width=150, height=100, corner_radius=15)
                    admin_frame.pack()
                    admin_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

                    def settings():
                        settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                        settings_frame.pack(fill="both", expand=True)

                        settings_1 = ctk.CTkFrame(master=settings_frame)
                        settings_1.pack(pady=20, padx=60, fill="both", expand=True)

                        settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                        settings_label.pack(padx=10, pady=12)
                        settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                        def change_appearance_mode_event(new_appearance_mode: str):
                            ctk.set_appearance_mode(new_appearance_mode)
                            fa = open("appearance.txt", "w")
                            fa.write(new_appearance_mode)
                            fa.close()

                        appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:")
                        appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                        appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                                   command=change_appearance_mode_event)
                        appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                        def change_scaling_event(new_scaling: str):
                            new_scaling_float = int(new_scaling.replace("%", "")) / 100
                            ctk.set_widget_scaling(new_scaling_float)
                            fs = open('Scale.txt', 'w')
                            n_scaling = new_scaling.replace("%", "")
                            fs.write(n_scaling)
                            fs.close()

                        scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:")
                        scaling_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)
                        scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                           command=change_scaling_event)
                        scaling_option.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)

                        def destroy_settings():
                            settings_frame.destroy()
                            settings_1.destroy()
                            settings_label.destroy()
                            appearance_mode_label.destroy()
                            appearance_mode_option.destroy()
                            scaling_label.destroy()
                            scaling_option.destroy()
                            _save.destroy()

                        _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15),
                                              command=destroy_settings)
                        _save.pack(padx=10, pady=12)
                        _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

                    settings_button = ctk.CTkButton(master=admin_frame, text="Settings", font=('Segoe Ui', 15),
                                                    command=settings)
                    settings_button.pack(padx=10, pady=12)
                    settings_button.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

                    _quit = ctk.CTkButton(master=admin_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
                    _quit.pack(padx=10, pady=12)
                    _quit.place(relx=0.5, rely=0.7, anchor=ctk.CENTER)
                else:
                    admin_frame.destroy()

            admin2_checkbox = ctk.CTkCheckBox(master=frame, text='Admin', corner_radius=5, font=('Segoe Ui', 21),
                                              variable=acb_va2, onvalue=1, offvalue=0, command=account2)
            admin2_checkbox.pack()
            admin2_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

            user2_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

            admin2_label = ctk.CTkLabel(master=frame, text="", image=user2_image)
            admin2_label.pack()
            admin2_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

            path = "timetable_teachers.xlsx"
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook['Class 11 A']
            header = sheet.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            body = sheet.iter_rows(min_row=2, max_row=6, max_col=9, values_only=True)
            header = [r for r in header]
            body = [r for r in body]
            workbook.close()

            style = ttk.Style()
            style.configure("Treeview", rowheight=60)
            style.configure("Treeview.Heading", font=('Segoe UI', 20))

            tree = ttk.Treeview(master=frame, selectmode='browse', show='headings', columns=header[0], height=5)
            tree.pack(padx=10, pady=10, fill='x')
            tree.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            tree.tag_configure("tag_item", font=('Segoe Ui', 12), foreground='black')

            for i in header[0]:
                tree.column(i, anchor=ctk.CENTER)
                tree.heading(i, text=i)

            for i in body:
                tree.insert('', 'end', iid=i[0], values=i, tags="tag_item")

            text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
            text_name.pack(padx=1, pady=2)
            text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

            def _back():
                view1_label.destroy()
                view1_back.destroy()
                text_name.destroy()
                class_label.destroy()
                tree.destroy()
                admin2_checkbox.destroy()
                admin2_label.destroy()
                if acb_va2.get() == 1:
                    admin_frame.destroy()
                class11()

            view1_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
            view1_back.pack(pady=60)
            view1_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

        def view2():

            view2_label = ctk.CTkLabel(master=frame, text="Timetables", font=('Segoe Ui', 40))
            view2_label.pack()
            view2_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

            class_label = ctk.CTkLabel(master=frame, text="Class 11 C", font=('Segoe Ui', 35))
            class_label.pack()
            class_label.place(relx=0.5, rely=0.2, anchor=ctk.N)

            acb_va2 = ctk.IntVar(value=0)

            # noinspection PyUnboundLocalVariable,PyGlobalUndefined
            def account2():
                if acb_va2.get() == 1:

                    global admin_frame

                    admin_frame = ctk.CTkFrame(master=frame, width=150, height=100, corner_radius=15)
                    admin_frame.pack()
                    admin_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

                    def settings():
                        settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                        settings_frame.pack(fill="both", expand=True)

                        settings_1 = ctk.CTkFrame(master=settings_frame)
                        settings_1.pack(pady=20, padx=60, fill="both", expand=True)

                        settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                        settings_label.pack(padx=10, pady=12)
                        settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                        def change_appearance_mode_event(new_appearance_mode: str):
                            ctk.set_appearance_mode(new_appearance_mode)
                            fa = open("appearance.txt", "w")
                            fa.write(new_appearance_mode)
                            fa.close()

                        appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:")
                        appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                        appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                                   command=change_appearance_mode_event)
                        appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                        def change_scaling_event(new_scaling: str):
                            new_scaling_float = int(new_scaling.replace("%", "")) / 100
                            ctk.set_widget_scaling(new_scaling_float)
                            fs = open('Scale.txt', 'w')
                            n_scaling = new_scaling.replace("%", "")
                            fs.write(n_scaling)
                            fs.close()

                        scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:")
                        scaling_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)
                        scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                           command=change_scaling_event)
                        scaling_option.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)

                        def destroy_settings():
                            settings_frame.destroy()
                            settings_1.destroy()
                            settings_label.destroy()
                            appearance_mode_label.destroy()
                            appearance_mode_option.destroy()
                            scaling_label.destroy()
                            scaling_option.destroy()
                            _save.destroy()

                        _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15),
                                              command=destroy_settings)
                        _save.pack(padx=10, pady=12)
                        _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

                    settings_button = ctk.CTkButton(master=admin_frame, text="Settings", font=('Segoe Ui', 15),
                                                    command=settings)
                    settings_button.pack(padx=10, pady=12)
                    settings_button.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

                    _quit = ctk.CTkButton(master=admin_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
                    _quit.pack(padx=10, pady=12)
                    _quit.place(relx=0.5, rely=0.7, anchor=ctk.CENTER)
                else:
                    admin_frame.destroy()

            admin2_checkbox = ctk.CTkCheckBox(master=frame, text='Admin', corner_radius=5, font=('Segoe Ui', 21),
                                              variable=acb_va2, onvalue=1, offvalue=0, command=account2)
            admin2_checkbox.pack()
            admin2_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

            user2_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

            admin2_label = ctk.CTkLabel(master=frame, text="", image=user2_image)
            admin2_label.pack()
            admin2_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

            path = "timetable_teachers.xlsx"
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook['Class 11 B']
            header = sheet.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            body = sheet.iter_rows(min_row=2, max_row=6, max_col=9, values_only=True)
            header = [r for r in header]
            body = [r for r in body]
            workbook.close()

            style = ttk.Style()
            style.configure("Treeview", rowheight=60)
            style.configure("Treeview.Heading", font=('Segoe UI', 20))

            tree = ttk.Treeview(master=frame, selectmode='browse', show='headings', columns=header[0], height=5)
            tree.pack(padx=10, pady=10, fill='x')
            tree.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            tree.tag_configure("tag_item", font=('Segoe Ui', 12), foreground='black')

            for i in header[0]:
                tree.column(i, anchor=ctk.CENTER)
                tree.heading(i, text=i)

            for i in body:
                tree.insert('', 'end', iid=i[0], values=i, tags="tag_item")

            text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
            text_name.pack(padx=1, pady=2)
            text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

            def _back():
                view2_label.destroy()
                view2_back.destroy()
                text_name.destroy()
                class_label.destroy()
                tree.destroy()
                admin2_checkbox.destroy()
                admin2_label.destroy()
                if acb_va2.get() == 1:
                    admin_frame.destroy()
                class11()

            view2_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
            view2_back.pack(pady=60)
            view2_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

        def go_view1():
            class10_label.destroy()
            class10_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class10_name.destroy()
            class10_back.destroy()
            admin1_checkbox.destroy()
            admin1_label.destroy()
            if acb_va1.get() == 1:
                admin_frame.destroy()
            view1()

        def go_view2():
            class10_label.destroy()
            class10_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class10_name.destroy()
            class10_back.destroy()
            admin1_label.destroy()
            admin1_checkbox.destroy()
            if acb_va1.get() == 1:
                admin_frame.destroy()
            view2()

        view1_button = ctk.CTkButton(frame, text="Class 10 A", font=('Segoe Ui', 25), command=go_view1)
        view1_button.pack(padx=10, pady=30)
        view1_button.place(relx=0.5, rely=0.45, anchor=ctk.N)

        view2_button = ctk.CTkButton(frame, text="Class 10 C", font=('Segoe Ui', 25), command=go_view2)
        view2_button.pack(padx=10, pady=30)
        view2_button.place(relx=0.5, rely=0.55, anchor=ctk.N)

        def back10():
            class10_label.destroy()
            class10_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class10_name.destroy()
            class10_back.destroy()
            admin1_checkbox.destroy()
            admin1_label.destroy()
            if acb_va1.get() == 1:
                admin_frame.destroy()
            tabview()

        class10_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
        class10_name.pack(padx=1, pady=2)
        class10_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

        class10_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=back10)
        class10_back.pack(pady=60)
        class10_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

    def go_class10():
        view_label.destroy()
        view_label2.destroy()
        view_back.destroy()
        class10_button.destroy()
        class11_button.destroy()
        admin_checkbox.destroy()
        admin_label.destroy()
        _name.destroy()
        if acb_va.get() == 1:
            admin_frame.destroy()
        class10()

    def go_class11():
        view_label.destroy()
        view_label2.destroy()
        view_back.destroy()
        class10_button.destroy()
        class11_button.destroy()
        admin_label.destroy()
        admin_checkbox.destroy()
        _name.destroy()
        if acb_va.get() == 1:
            admin_frame.destroy()
        class11()

    class10_button = ctk.CTkButton(frame, text="Class 10", font=('Segoe Ui', 25), command=go_class10)
    class10_button.pack(padx=10, pady=30)
    class10_button.place(relx=0.5, rely=0.45, anchor=ctk.N)

    class11_button = ctk.CTkButton(frame, text="Class 11", font=('Segoe Ui', 25), command=go_class11)
    class11_button.pack(padx=10, pady=30)
    class11_button.place(relx=0.5, rely=0.55, anchor=ctk.N)

    _name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
    _name.pack(padx=1, pady=2)
    _name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

    def back():
        view_label.destroy()
        view_label2.destroy()
        view_back.destroy()
        class10_button.destroy()
        class11_button.destroy()
        admin_label.destroy()
        admin_checkbox.destroy()
        _name.destroy()
        if acb_va.get() == 1:
            admin_frame.destroy()
        tab2()

    view_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=back)
    view_back.pack(pady=60)
    view_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)


# to manage timetables - Admin
def tabman():
    man_label = ctk.CTkLabel(master=frame, text="Manage Timetables", font=('Segoe Ui', 40))
    man_label.pack(padx=20, pady=100)
    man_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

    acb_va = ctk.IntVar(value=0)

    # noinspection PyUnboundLocalVariable,PyGlobalUndefined
    def account():
        if acb_va.get() == 1:

            global admin_frame

            admin_frame = ctk.CTkFrame(master=frame, width=150, height=100, corner_radius=15)
            admin_frame.pack()
            admin_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

            def settings():
                settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                settings_frame.pack(fill="both", expand=True)

                settings_1 = ctk.CTkFrame(master=settings_frame, corner_radius=30)
                settings_1.pack(pady=60, padx=550, fill="both", expand=True)

                settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                settings_label.pack(padx=10, pady=12)
                settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                def change_appearance_mode_event(new_appearance_mode: str):
                    ctk.set_appearance_mode(new_appearance_mode)
                    fa = open("appearance.txt", "w")
                    fa.write(new_appearance_mode)
                    fa.close()

                appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:", font=('Segoe Ui', 20))
                appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                           command=change_appearance_mode_event, font=('Segoe Ui', 15))
                appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                def change_scaling_event(new_scaling: str):
                    new_scaling_float = int(new_scaling.replace("%", "")) / 100
                    ctk.set_widget_scaling(new_scaling_float)
                    fs = open('Scale.txt', 'w')
                    n_scaling = new_scaling.replace("%", "")
                    fs.write(n_scaling)
                    fs.close()

                scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:", font=('Segoe Ui', 20))
                scaling_label.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)
                scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                   command=change_scaling_event, font=('Segoe Ui', 15))
                scaling_option.place(relx=0.5, rely=0.6, anchor=ctk.CENTER)

                def destroy_settings():
                    settings_frame.destroy()
                    settings_1.destroy()
                    settings_label.destroy()
                    appearance_mode_label.destroy()
                    appearance_mode_option.destroy()
                    scaling_label.destroy()
                    scaling_option.destroy()
                    _save.destroy()

                _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15), command=destroy_settings)
                _save.pack(padx=10, pady=12)
                _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

            settings_button = ctk.CTkButton(master=admin_frame, text="Settings", font=('Segoe Ui', 15),
                                            command=settings)
            settings_button.pack(padx=10, pady=12)
            settings_button.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

            _quit = ctk.CTkButton(master=admin_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
            _quit.pack(padx=10, pady=12)
            _quit.place(relx=0.5, rely=0.7, anchor=ctk.CENTER)
        else:
            admin_frame.destroy()

    admin_checkbox = ctk.CTkCheckBox(master=frame, text='Admin', corner_radius=5, font=('Segoe Ui', 21),
                                     variable=acb_va, onvalue=1, offvalue=0, command=account)
    admin_checkbox.pack()
    admin_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

    user_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

    admin_label = ctk.CTkLabel(master=frame, text="", image=user_image)
    admin_label.pack()
    admin_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

    wip_label = ctk.CTkLabel(master=frame, text="Work in progress", font=('Segoe Ui', 35))
    wip_label.pack(padx=20, pady=100)
    wip_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

    wip_label2 = ctk.CTkLabel(master=frame, text="will be added later on...", font=('Segoe Ui', 20))
    wip_label2.pack(padx=20, pady=10)
    wip_label2.place(relx=0.5, rely=0.58, anchor=ctk.CENTER)

    text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
    text_name.pack(padx=1, pady=2)
    text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

    def _back():
        man_label.destroy()
        man_back.destroy()
        admin_label.destroy()
        admin_checkbox.destroy()
        wip_label.destroy()
        wip_label2.destroy()
        text_name.destroy()
        if acb_va.get() == 1:
            admin_frame.destroy()
        tab2()

    man_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
    man_back.pack(pady=60)
    man_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)


# to substitute absent teachers
def tabsub():
    sub_label = ctk.CTkLabel(master=frame, text="Substitute Absent Teachers", font=('Segoe Ui', 40))
    sub_label.pack(padx=20, pady=100)
    sub_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

    acb_va = ctk.IntVar(value=0)

    # noinspection PyUnboundLocalVariable,PyGlobalUndefined
    def account():
        if acb_va.get() == 1:

            global admin_frame

            admin_frame = ctk.CTkFrame(master=frame, width=150, height=100, corner_radius=15)
            admin_frame.pack()
            admin_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

            def settings():
                settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                settings_frame.pack(fill="both", expand=True)

                settings_1 = ctk.CTkFrame(master=settings_frame, corner_radius=30)
                settings_1.pack(pady=60, padx=550, fill="both", expand=True)

                settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                settings_label.pack(padx=10, pady=12)
                settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                def change_appearance_mode_event(new_appearance_mode: str):
                    ctk.set_appearance_mode(new_appearance_mode)
                    fa = open("appearance.txt", "w")
                    fa.write(new_appearance_mode)
                    fa.close()

                appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:", font=('Segoe Ui', 20))
                appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                           command=change_appearance_mode_event, font=('Segoe Ui', 15))
                appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                def change_scaling_event(new_scaling: str):
                    new_scaling_float = int(new_scaling.replace("%", "")) / 100
                    ctk.set_widget_scaling(new_scaling_float)
                    fs = open('Scale.txt', 'w')
                    n_scaling = new_scaling.replace("%", "")
                    fs.write(n_scaling)
                    fs.close()

                scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:", font=('Segoe Ui', 20))
                scaling_label.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)
                scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                   command=change_scaling_event, font=('Segoe Ui', 15))
                scaling_option.place(relx=0.5, rely=0.6, anchor=ctk.CENTER)

                def destroy_settings():
                    settings_frame.destroy()
                    settings_1.destroy()
                    settings_label.destroy()
                    appearance_mode_label.destroy()
                    appearance_mode_option.destroy()
                    scaling_label.destroy()
                    scaling_option.destroy()
                    _save.destroy()

                _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15), command=destroy_settings)
                _save.pack(padx=10, pady=12)
                _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

            settings_button = ctk.CTkButton(master=admin_frame, text="Settings", font=('Segoe Ui', 15),
                                            command=settings)
            settings_button.pack(padx=10, pady=12)
            settings_button.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

            _quit = ctk.CTkButton(master=admin_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
            _quit.pack(padx=10, pady=12)
            _quit.place(relx=0.5, rely=0.7, anchor=ctk.CENTER)
        else:
            admin_frame.destroy()

    admin_checkbox = ctk.CTkCheckBox(master=frame, text='Admin', corner_radius=5, font=('Segoe Ui', 21),
                                     variable=acb_va, onvalue=1, offvalue=0, command=account)
    admin_checkbox.pack()
    admin_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

    user_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

    admin_label = ctk.CTkLabel(master=frame, text="", image=user_image)
    admin_label.pack()
    admin_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

    wip_label = ctk.CTkLabel(master=frame, text="Work in progress", font=('Segoe Ui', 35))
    wip_label.pack(padx=20, pady=100)
    wip_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

    text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
    text_name.pack(padx=1, pady=2)
    text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

    def _back():
        sub_label.destroy()
        sub_back.destroy()
        admin_label.destroy()
        admin_checkbox.destroy()
        wip_label.destroy()
        text_name.destroy()
        if acb_va.get() == 1:
            admin_frame.destroy()
        tab2()

    sub_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
    sub_back.pack(pady=60)
    sub_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)


# program lobby - Teachers
# if username and password is entered for teachers it runs this code
def ttab2():
    def destroy_ttab2():
        label1.destroy()
        view_button.destroy()
        sub_button.destroy()
        text_name.destroy()
        teach_label.destroy()
        teach_checkbox.destroy()

    label1 = ctk.CTkLabel(master=frame, text="Time Management System", font=('Segoe Ui', 55))
    label1.pack(padx=10, pady=12)
    label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

    acb_va = ctk.IntVar(value=0)

    # noinspection PyUnboundLocalVariable,PyGlobalUndefined
    def account():
        if acb_va.get() == 1:

            global admin_frame

            admin_frame = ctk.CTkFrame(master=frame, width=150, height=150, corner_radius=15)
            admin_frame.pack()
            admin_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

            def settings():
                settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                settings_frame.pack(fill="both", expand=True)

                settings_1 = ctk.CTkFrame(master=settings_frame, corner_radius=30)
                settings_1.pack(pady=60, padx=550, fill="both", expand=True)

                settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                settings_label.pack(padx=10, pady=12)
                settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                def change_appearance_mode_event(new_appearance_mode: str):
                    ctk.set_appearance_mode(new_appearance_mode)
                    fa = open("appearance.txt", "w")
                    fa.write(new_appearance_mode)
                    fa.close()

                appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:", font=('Segoe Ui', 20))
                appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                           command=change_appearance_mode_event, font=('Segoe Ui', 15))
                appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                def change_scaling_event(new_scaling: str):
                    new_scaling_float = int(new_scaling.replace("%", "")) / 100
                    ctk.set_widget_scaling(new_scaling_float)
                    fs = open('Scale.txt', 'w')
                    n_scaling = new_scaling.replace("%", "")
                    fs.write(n_scaling)
                    fs.close()

                scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:", font=('Segoe Ui', 20))
                scaling_label.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)
                scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                   command=change_scaling_event, font=('Segoe Ui', 15))
                scaling_option.place(relx=0.5, rely=0.6, anchor=ctk.CENTER)

                def destroy_settings():
                    settings_frame.destroy()
                    settings_1.destroy()
                    settings_label.destroy()
                    appearance_mode_label.destroy()
                    appearance_mode_option.destroy()
                    scaling_label.destroy()
                    scaling_option.destroy()
                    _save.destroy()

                _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15), command=destroy_settings)
                _save.pack(padx=10, pady=12)
                _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

            def logout():
                fh = open('login remember teach.txt', 'w')
                str2 = "false"
                print(str2)
                fh.write(str2)
                fh.close()
                destroy_account()
                destroy_ttab2a()
                tab1()

            def destroy_account():
                admin_frame.destroy()
                settings_button.destroy()
                logout1.destroy()
                _quit.destroy()

            def destroy_ttab2a():
                label1.destroy()
                view_button.destroy()
                sub_button.destroy()
                teach_label.destroy()
                teach_checkbox.destroy()
                logout1.destroy()
                text_name.destroy()
                _quit.destroy()

            settings_button = ctk.CTkButton(master=admin_frame, text="Settings", font=('Segoe Ui', 15),
                                            command=settings)
            settings_button.pack(padx=10, pady=12)
            settings_button.place(relx=0.5, rely=0.25, anchor=ctk.CENTER)

            logout1 = ctk.CTkButton(master=admin_frame, text="Logout", font=('Segoe Ui', 15), command=logout)
            logout1.pack(padx=10, pady=10)
            logout1.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            _quit = ctk.CTkButton(master=admin_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
            _quit.pack(padx=10, pady=12)
            _quit.place(relx=0.5, rely=0.75, anchor=ctk.CENTER)
        else:
            admin_frame.destroy()

    teach_checkbox = ctk.CTkCheckBox(master=frame, text='Teacher', corner_radius=5, font=('Segoe Ui', 21),
                                     variable=acb_va, onvalue=1, offvalue=0, command=account)
    teach_checkbox.pack()
    teach_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

    user_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

    teach_label = ctk.CTkLabel(master=frame, text="", image=user_image)
    teach_label.pack()
    teach_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

    def button_view():
        destroy_ttab2()
        ttabview()

    def button_sub():
        destroy_ttab2()
        ttabsub()

    view_button = ctk.CTkButton(frame, text="View Timetables", font=('Segoe Ui', 45), command=button_view)
    view_button.pack(padx=10, pady=30)
    view_button.place(relx=0.5, rely=0.43, anchor=ctk.N)

    sub_button = ctk.CTkButton(frame, text="View Today's Timetables", font=('Segoe Ui', 45), command=button_sub)
    sub_button.pack(padx=10, pady=30)
    sub_button.place(relx=0.5, rely=0.58, anchor=ctk.N)

    text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
    text_name.pack(padx=1, pady=2)
    text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)


# to view the timetable - Teachers
def ttabview():
    view_label = ctk.CTkLabel(master=frame, text="Timetables", font=('Segoe Ui', 40))
    view_label.pack()
    view_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

    view_label2 = ctk.CTkLabel(master=frame, text="Pick a class", font=('Segoe Ui', 35))
    view_label2.pack()
    view_label2.place(relx=0.5, rely=0.2, anchor=ctk.N)

    acb_va = ctk.IntVar(value=0)

    # noinspection PyUnboundLocalVariable,PyGlobalUndefined
    def account():
        if acb_va.get() == 1:

            global teach_frame

            teach_frame = ctk.CTkFrame(master=frame, width=150, height=100, corner_radius=15)
            teach_frame.pack()
            teach_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

            def settings():
                settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                settings_frame.pack(fill="both", expand=True)

                settings_1 = ctk.CTkFrame(master=settings_frame, corner_radius=30)
                settings_1.pack(pady=60, padx=550, fill="both", expand=True)

                settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                settings_label.pack(padx=10, pady=12)
                settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                def change_appearance_mode_event(new_appearance_mode: str):
                    ctk.set_appearance_mode(new_appearance_mode)
                    fa = open("appearance.txt", "w")
                    fa.write(new_appearance_mode)
                    fa.close()

                appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:", font=('Segoe Ui', 20))
                appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                           command=change_appearance_mode_event, font=('Segoe Ui', 15))
                appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                def change_scaling_event(new_scaling: str):
                    new_scaling_float = int(new_scaling.replace("%", "")) / 100
                    ctk.set_widget_scaling(new_scaling_float)
                    fs = open('Scale.txt', 'w')
                    n_scaling = new_scaling.replace("%", "")
                    fs.write(n_scaling)
                    fs.close()

                scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:", font=('Segoe Ui', 20))
                scaling_label.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)
                scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                   command=change_scaling_event, font=('Segoe Ui', 15))
                scaling_option.place(relx=0.5, rely=0.6, anchor=ctk.CENTER)

                def destroy_settings():
                    settings_frame.destroy()
                    settings_1.destroy()
                    settings_label.destroy()
                    appearance_mode_label.destroy()
                    appearance_mode_option.destroy()
                    scaling_label.destroy()
                    scaling_option.destroy()
                    _save.destroy()

                _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15), command=destroy_settings)
                _save.pack(padx=10, pady=12)
                _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

            settings_button = ctk.CTkButton(master=teach_frame, text="Settings", font=('Segoe Ui', 15),
                                            command=settings)
            settings_button.pack(padx=10, pady=12)
            settings_button.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

            _quit = ctk.CTkButton(master=teach_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
            _quit.pack(padx=10, pady=12)
            _quit.place(relx=0.5, rely=0.7, anchor=ctk.CENTER)
        else:
            teach_frame.destroy()

    teach_checkbox = ctk.CTkCheckBox(master=frame, text='Teacher', corner_radius=5, font=('Segoe Ui', 21),
                                     variable=acb_va, onvalue=1, offvalue=0, command=account)
    teach_checkbox.pack()
    teach_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

    user_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

    teach_label = ctk.CTkLabel(master=frame, text="", image=user_image)
    teach_label.pack()
    teach_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

    # reading the Excel files

    def class11():

        class11_label = ctk.CTkLabel(master=frame, text="Timetables", font=('Segoe Ui', 40))
        class11_label.pack()
        class11_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

        class11_label2 = ctk.CTkLabel(master=frame, text="Class 11", font=('Segoe Ui', 35))
        class11_label2.pack()
        class11_label2.place(relx=0.5, rely=0.2, anchor=ctk.N)

        acb_va1 = ctk.IntVar(value=0)

        # noinspection PyUnboundLocalVariable,PyGlobalUndefined
        def account1():
            if acb_va1.get() == 1:

                global teach_frame

                teach_frame = ctk.CTkFrame(master=frame, width=150, height=100, corner_radius=15)
                teach_frame.pack()
                teach_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

                def settings():
                    settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                    settings_frame.pack(fill="both", expand=True)

                    settings_1 = ctk.CTkFrame(master=settings_frame)
                    settings_1.pack(pady=20, padx=60, fill="both", expand=True)

                    settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                    settings_label.pack(padx=10, pady=12)
                    settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                    def change_appearance_mode_event(new_appearance_mode: str):
                        ctk.set_appearance_mode(new_appearance_mode)
                        fa = open("appearance.txt", "w")
                        fa.write(new_appearance_mode)
                        fa.close()

                    appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:")
                    appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                    appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                               command=change_appearance_mode_event)
                    appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                    def change_scaling_event(new_scaling: str):
                        new_scaling_float = int(new_scaling.replace("%", "")) / 100
                        ctk.set_widget_scaling(new_scaling_float)
                        fs = open('Scale.txt', 'w')
                        n_scaling = new_scaling.replace("%", "")
                        fs.write(n_scaling)
                        fs.close()

                    scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:")
                    scaling_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)
                    scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                       command=change_scaling_event)
                    scaling_option.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)

                    def destroy_settings():
                        settings_frame.destroy()
                        settings_1.destroy()
                        settings_label.destroy()
                        appearance_mode_label.destroy()
                        appearance_mode_option.destroy()
                        scaling_label.destroy()
                        scaling_option.destroy()
                        _save.destroy()

                    _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15),
                                          command=destroy_settings)
                    _save.pack(padx=10, pady=12)
                    _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

                settings_button = ctk.CTkButton(master=teach_frame, text="Settings", font=('Segoe Ui', 15),
                                                command=settings)
                settings_button.pack(padx=10, pady=12)
                settings_button.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

                _quit = ctk.CTkButton(master=teach_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
                _quit.pack(padx=10, pady=12)
                _quit.place(relx=0.5, rely=0.7, anchor=ctk.CENTER)
            else:
                teach_frame.destroy()

        teach1_checkbox = ctk.CTkCheckBox(master=frame, text='Teacher', corner_radius=5, font=('Segoe Ui', 21),
                                          variable=acb_va1, onvalue=1, offvalue=0, command=account1)
        teach1_checkbox.pack()
        teach1_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

        user1_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

        teach1_label = ctk.CTkLabel(master=frame, text="", image=user1_image)
        teach1_label.pack()
        teach1_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

        def view1():

            view1_label = ctk.CTkLabel(master=frame, text="Timetables", font=('Segoe Ui', 40))
            view1_label.pack()
            view1_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

            class_label = ctk.CTkLabel(master=frame, text="Class 11 A", font=('Segoe Ui', 35))
            class_label.pack()
            class_label.place(relx=0.5, rely=0.2, anchor=ctk.N)

            acb_va2 = ctk.IntVar(value=0)

            # noinspection PyUnboundLocalVariable,PyGlobalUndefined
            def account2():
                if acb_va2.get() == 1:

                    global teach_frame

                    teach_frame = ctk.CTkFrame(master=frame, width=150, height=100, corner_radius=15)
                    teach_frame.pack()
                    teach_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

                    def settings():
                        settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                        settings_frame.pack(fill="both", expand=True)

                        settings_1 = ctk.CTkFrame(master=settings_frame)
                        settings_1.pack(pady=20, padx=60, fill="both", expand=True)

                        settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                        settings_label.pack(padx=10, pady=12)
                        settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                        def change_appearance_mode_event(new_appearance_mode: str):
                            ctk.set_appearance_mode(new_appearance_mode)
                            fa = open("appearance.txt", "w")
                            fa.write(new_appearance_mode)
                            fa.close()

                        appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:")
                        appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                        appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                                   command=change_appearance_mode_event)
                        appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                        def change_scaling_event(new_scaling: str):
                            new_scaling_float = int(new_scaling.replace("%", "")) / 100
                            ctk.set_widget_scaling(new_scaling_float)
                            fs = open('Scale.txt', 'w')
                            n_scaling = new_scaling.replace("%", "")
                            fs.write(n_scaling)
                            fs.close()

                        scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:")
                        scaling_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)
                        scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                           command=change_scaling_event)
                        scaling_option.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)

                        def destroy_settings():
                            settings_frame.destroy()
                            settings_1.destroy()
                            settings_label.destroy()
                            appearance_mode_label.destroy()
                            appearance_mode_option.destroy()
                            scaling_label.destroy()
                            scaling_option.destroy()
                            _save.destroy()

                        _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15),
                                              command=destroy_settings)
                        _save.pack(padx=10, pady=12)
                        _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

                    settings_button = ctk.CTkButton(master=teach_frame, text="Settings", font=('Segoe Ui', 15),
                                                    command=settings)
                    settings_button.pack(padx=10, pady=12)
                    settings_button.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

                    _quit = ctk.CTkButton(master=teach_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
                    _quit.pack(padx=10, pady=12)
                    _quit.place(relx=0.5, rely=0.7, anchor=ctk.CENTER)
                else:
                    teach_frame.destroy()

            teach2_checkbox = ctk.CTkCheckBox(master=frame, text='Teacher', corner_radius=5, font=('Segoe Ui', 21),
                                              variable=acb_va2, onvalue=1, offvalue=0, command=account2)
            teach2_checkbox.pack()
            teach2_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

            user2_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

            teach2_label = ctk.CTkLabel(master=frame, text="", image=user2_image)
            teach2_label.pack()
            teach2_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

            path = "timetable_teachers.xlsx"
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook['Class 11 A']
            header = sheet.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            body = sheet.iter_rows(min_row=2, max_row=6, max_col=9, values_only=True)
            header = [r for r in header]
            body = [r for r in body]
            workbook.close()

            style = ttk.Style()
            style.configure("Treeview", rowheight=60)
            style.configure("Treeview.Heading", font=('Segoe UI', 20))

            tree = ttk.Treeview(master=frame, selectmode='browse', show='headings', columns=header[0], height=5)
            tree.pack(padx=10, pady=10, fill='x')
            tree.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            tree.tag_configure("tag_item", font=('Segoe Ui', 12), foreground='black')

            for i in header[0]:
                tree.column(i, anchor=ctk.CENTER)
                tree.heading(i, text=i)

            for i in body:
                tree.insert('', 'end', iid=i[0], values=i, tags="tag_item")

            text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
            text_name.pack(padx=1, pady=2)
            text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

            def _back():
                view1_label.destroy()
                view1_back.destroy()
                text_name.destroy()
                class_label.destroy()
                tree.destroy()
                teach2_checkbox.destroy()
                teach2_label.destroy()
                if acb_va2.get() == 1:
                    teach_frame.destroy()
                class11()

            view1_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
            view1_back.pack(pady=60)
            view1_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

        def view2():

            view2_label = ctk.CTkLabel(master=frame, text="Timetables", font=('Segoe Ui', 40))
            view2_label.pack()
            view2_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

            class_label = ctk.CTkLabel(master=frame, text="Class 11 C", font=('Segoe Ui', 35))
            class_label.pack()
            class_label.place(relx=0.5, rely=0.2, anchor=ctk.N)

            acb_va2 = ctk.IntVar(value=0)

            # noinspection PyUnboundLocalVariable,PyGlobalUndefined
            def account2():
                if acb_va2.get() == 1:

                    global teach_frame

                    teach_frame = ctk.CTkFrame(master=frame, width=150, height=100, corner_radius=15)
                    teach_frame.pack()
                    teach_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

                    def settings():
                        settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                        settings_frame.pack(fill="both", expand=True)

                        settings_1 = ctk.CTkFrame(master=settings_frame)
                        settings_1.pack(pady=20, padx=60, fill="both", expand=True)

                        settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                        settings_label.pack(padx=10, pady=12)
                        settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                        def change_appearance_mode_event(new_appearance_mode: str):
                            ctk.set_appearance_mode(new_appearance_mode)
                            fa = open("appearance.txt", "w")
                            fa.write(new_appearance_mode)
                            fa.close()

                        appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:")
                        appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                        appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                                   command=change_appearance_mode_event)
                        appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                        def change_scaling_event(new_scaling: str):
                            new_scaling_float = int(new_scaling.replace("%", "")) / 100
                            ctk.set_widget_scaling(new_scaling_float)
                            fs = open('Scale.txt', 'w')
                            n_scaling = new_scaling.replace("%", "")
                            fs.write(n_scaling)
                            fs.close()

                        scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:")
                        scaling_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)
                        scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                           command=change_scaling_event)
                        scaling_option.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)

                        def destroy_settings():
                            settings_frame.destroy()
                            settings_1.destroy()
                            settings_label.destroy()
                            appearance_mode_label.destroy()
                            appearance_mode_option.destroy()
                            scaling_label.destroy()
                            scaling_option.destroy()
                            _save.destroy()

                        _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15),
                                              command=destroy_settings)
                        _save.pack(padx=10, pady=12)
                        _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

                    settings_button = ctk.CTkButton(master=teach_frame, text="Settings", font=('Segoe Ui', 15),
                                                    command=settings)
                    settings_button.pack(padx=10, pady=12)
                    settings_button.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

                    _quit = ctk.CTkButton(master=teach_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
                    _quit.pack(padx=10, pady=12)
                    _quit.place(relx=0.5, rely=0.7, anchor=ctk.CENTER)
                else:
                    teach_frame.destroy()

            teach2_checkbox = ctk.CTkCheckBox(master=frame, text='Teacher', corner_radius=5, font=('Segoe Ui', 21),
                                              variable=acb_va2, onvalue=1, offvalue=0, command=account2)
            teach2_checkbox.pack()
            teach2_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

            user2_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

            teach2_label = ctk.CTkLabel(master=frame, text="", image=user2_image)
            teach2_label.pack()
            teach2_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

            path = "timetable_teachers.xlsx"
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook['Class 11 B']
            header = sheet.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            body = sheet.iter_rows(min_row=2, max_row=6, max_col=9, values_only=True)
            header = [r for r in header]
            body = [r for r in body]
            workbook.close()

            style = ttk.Style()
            style.configure("Treeview", rowheight=60)
            style.configure("Treeview.Heading", font=('Segoe UI', 20))

            tree = ttk.Treeview(master=frame, selectmode='browse', show='headings', columns=header[0], height=5)
            tree.pack(padx=10, pady=10, fill='x')
            tree.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            tree.tag_configure("tag_item", font=('Segoe Ui', 12), foreground='black')

            for i in header[0]:
                tree.column(i, anchor=ctk.CENTER)
                tree.heading(i, text=i)

            for i in body:
                tree.insert('', 'end', iid=i[0], values=i, tags="tag_item")

            text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
            text_name.pack(padx=1, pady=2)
            text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

            def _back():
                view2_label.destroy()
                view2_back.destroy()
                text_name.destroy()
                class_label.destroy()
                tree.destroy()
                teach2_checkbox.destroy()
                teach2_label.destroy()
                if acb_va2.get() == 1:
                    teach_frame.destroy()
                class11()

            view2_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
            view2_back.pack(pady=60)
            view2_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

        def go_view1():
            class11_label.destroy()
            class11_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class11_name.destroy()
            class11_back.destroy()
            teach1_checkbox.destroy()
            teach1_label.destroy()
            if acb_va1.get() == 1:
                teach_frame.destroy()
            view1()

        def go_view2():
            class11_label.destroy()
            class11_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class11_name.destroy()
            class11_back.destroy()
            teach1_label.destroy()
            teach1_checkbox.destroy()
            if acb_va1.get() == 1:
                teach_frame.destroy()
            view2()

        view1_button = ctk.CTkButton(frame, text="Class 11 A", font=('Segoe Ui', 25), command=go_view1)
        view1_button.pack(padx=10, pady=30)
        view1_button.place(relx=0.5, rely=0.45, anchor=ctk.N)

        view2_button = ctk.CTkButton(frame, text="Class 11 C", font=('Segoe Ui', 25), command=go_view2)
        view2_button.pack(padx=10, pady=30)
        view2_button.place(relx=0.5, rely=0.55, anchor=ctk.N)

        def back11():
            class11_label.destroy()
            class11_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class11_name.destroy()
            class11_back.destroy()
            teach1_label.destroy()
            teach1_checkbox.destroy()
            if acb_va1.get() == 1:
                teach_frame.destroy()
            tabview()

        class11_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
        class11_name.pack(padx=1, pady=2)
        class11_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

        class11_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=back11)
        class11_back.pack()
        class11_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

    def class10():

        class10_label = ctk.CTkLabel(master=frame, text="Timetables", font=('Segoe Ui', 40))
        class10_label.pack()
        class10_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

        class10_label2 = ctk.CTkLabel(master=frame, text="Class 10", font=('Segoe Ui', 35))
        class10_label2.pack()
        class10_label2.place(relx=0.5, rely=0.2, anchor=ctk.N)

        acb_va1 = ctk.IntVar(value=0)

        # noinspection PyUnboundLocalVariable,PyGlobalUndefined
        def account1():
            if acb_va1.get() == 1:

                global teach_frame

                teach_frame = ctk.CTkFrame(master=frame, width=150, height=100, corner_radius=15)
                teach_frame.pack()
                teach_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

                def settings():
                    settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                    settings_frame.pack(fill="both", expand=True)

                    settings_1 = ctk.CTkFrame(master=settings_frame)
                    settings_1.pack(pady=20, padx=60, fill="both", expand=True)

                    settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                    settings_label.pack(padx=10, pady=12)
                    settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                    def change_appearance_mode_event(new_appearance_mode: str):
                        ctk.set_appearance_mode(new_appearance_mode)
                        fa = open("appearance.txt", "w")
                        fa.write(new_appearance_mode)
                        fa.close()

                    appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:")
                    appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                    appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                               command=change_appearance_mode_event)
                    appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                    def change_scaling_event(new_scaling: str):
                        new_scaling_float = int(new_scaling.replace("%", "")) / 100
                        ctk.set_widget_scaling(new_scaling_float)
                        fs = open('Scale.txt', 'w')
                        n_scaling = new_scaling.replace("%", "")
                        fs.write(n_scaling)
                        fs.close()

                    scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:")
                    scaling_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)
                    scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                       command=change_scaling_event)
                    scaling_option.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)

                    def destroy_settings():
                        settings_frame.destroy()
                        settings_1.destroy()
                        settings_label.destroy()
                        appearance_mode_label.destroy()
                        appearance_mode_option.destroy()
                        scaling_label.destroy()
                        scaling_option.destroy()
                        _save.destroy()

                    _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15),
                                          command=destroy_settings)
                    _save.pack(padx=10, pady=12)
                    _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

                settings_button = ctk.CTkButton(master=teach_frame, text="Settings", font=('Segoe Ui', 15),
                                                command=settings)
                settings_button.pack(padx=10, pady=12)
                settings_button.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

                _quit = ctk.CTkButton(master=teach_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
                _quit.pack(padx=10, pady=12)
                _quit.place(relx=0.5, rely=0.7, anchor=ctk.CENTER)
            else:
                teach_frame.destroy()

        teach1_checkbox = ctk.CTkCheckBox(master=frame, text='Teacher', corner_radius=5, font=('Segoe Ui', 21),
                                          variable=acb_va1, onvalue=1, offvalue=0, command=account1)
        teach1_checkbox.pack()
        teach1_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

        user1_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

        teach1_label = ctk.CTkLabel(master=frame, text="", image=user1_image)
        teach1_label.pack()
        teach1_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

        def view1():

            view1_label = ctk.CTkLabel(master=frame, text="Timetables", font=('Segoe Ui', 40))
            view1_label.pack()
            view1_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

            class_label = ctk.CTkLabel(master=frame, text="Class 11 A", font=('Segoe Ui', 35))
            class_label.pack()
            class_label.place(relx=0.5, rely=0.2, anchor=ctk.N)

            acb_va2 = ctk.IntVar(value=0)

            # noinspection PyUnboundLocalVariable,PyGlobalUndefined
            def account2():
                if acb_va2.get() == 1:

                    global teach_frame

                    teach_frame = ctk.CTkFrame(master=frame, width=150, height=100, corner_radius=15)
                    teach_frame.pack()
                    teach_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

                    def settings():
                        settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                        settings_frame.pack(fill="both", expand=True)

                        settings_1 = ctk.CTkFrame(master=settings_frame)
                        settings_1.pack(pady=20, padx=60, fill="both", expand=True)

                        settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                        settings_label.pack(padx=10, pady=12)
                        settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                        def change_appearance_mode_event(new_appearance_mode: str):
                            ctk.set_appearance_mode(new_appearance_mode)
                            fa = open("appearance.txt", "w")
                            fa.write(new_appearance_mode)
                            fa.close()

                        appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:")
                        appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                        appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                                   command=change_appearance_mode_event)
                        appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                        def change_scaling_event(new_scaling: str):
                            new_scaling_float = int(new_scaling.replace("%", "")) / 100
                            ctk.set_widget_scaling(new_scaling_float)
                            fs = open('Scale.txt', 'w')
                            n_scaling = new_scaling.replace("%", "")
                            fs.write(n_scaling)
                            fs.close()

                        scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:")
                        scaling_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)
                        scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                           command=change_scaling_event)
                        scaling_option.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)

                        def destroy_settings():
                            settings_frame.destroy()
                            settings_1.destroy()
                            settings_label.destroy()
                            appearance_mode_label.destroy()
                            appearance_mode_option.destroy()
                            scaling_label.destroy()
                            scaling_option.destroy()
                            _save.destroy()

                        _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15),
                                              command=destroy_settings)
                        _save.pack(padx=10, pady=12)
                        _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

                    settings_button = ctk.CTkButton(master=teach_frame, text="Settings", font=('Segoe Ui', 15),
                                                    command=settings)
                    settings_button.pack(padx=10, pady=12)
                    settings_button.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

                    _quit = ctk.CTkButton(master=teach_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
                    _quit.pack(padx=10, pady=12)
                    _quit.place(relx=0.5, rely=0.7, anchor=ctk.CENTER)
                else:
                    teach_frame.destroy()

            teach2_checkbox = ctk.CTkCheckBox(master=frame, text='Teacher', corner_radius=5, font=('Segoe Ui', 21),
                                              variable=acb_va2, onvalue=1, offvalue=0, command=account2)
            teach2_checkbox.pack()
            teach2_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

            user2_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

            teach2_label = ctk.CTkLabel(master=frame, text="", image=user2_image)
            teach2_label.pack()
            teach2_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

            path = "timetable_teachers.xlsx"
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook['Class 11 A']
            header = sheet.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            body = sheet.iter_rows(min_row=2, max_row=6, max_col=9, values_only=True)
            header = [r for r in header]
            body = [r for r in body]
            workbook.close()

            style = ttk.Style()
            style.configure("Treeview", rowheight=60)
            style.configure("Treeview.Heading", font=('Segoe UI', 20))

            tree = ttk.Treeview(master=frame, selectmode='browse', show='headings', columns=header[0], height=5)
            tree.pack(padx=10, pady=10, fill='x')
            tree.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            tree.tag_configure("tag_item", font=('Segoe Ui', 12), foreground='black')

            for i in header[0]:
                tree.column(i, anchor=ctk.CENTER)
                tree.heading(i, text=i)

            for i in body:
                tree.insert('', 'end', iid=i[0], values=i, tags="tag_item")

            text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
            text_name.pack(padx=1, pady=2)
            text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

            def _back():
                view1_label.destroy()
                view1_back.destroy()
                text_name.destroy()
                class_label.destroy()
                tree.destroy()
                teach2_checkbox.destroy()
                teach2_label.destroy()
                if acb_va2.get() == 1:
                    teach_frame.destroy()
                class11()

            view1_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
            view1_back.pack(pady=60)
            view1_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

        def view2():

            view2_label = ctk.CTkLabel(master=frame, text="Timetables", font=('Segoe Ui', 40))
            view2_label.pack()
            view2_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

            class_label = ctk.CTkLabel(master=frame, text="Class 11 C", font=('Segoe Ui', 35))
            class_label.pack()
            class_label.place(relx=0.5, rely=0.2, anchor=ctk.N)

            acb_va2 = ctk.IntVar(value=0)

            # noinspection PyUnboundLocalVariable,PyGlobalUndefined
            def account2():
                if acb_va2.get() == 1:

                    global teach_frame

                    teach_frame = ctk.CTkFrame(master=frame, width=150, height=100, corner_radius=15)
                    teach_frame.pack()
                    teach_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

                    def settings():
                        settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                        settings_frame.pack(fill="both", expand=True)

                        settings_1 = ctk.CTkFrame(master=settings_frame)
                        settings_1.pack(pady=20, padx=60, fill="both", expand=True)

                        settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                        settings_label.pack(padx=10, pady=12)
                        settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                        def change_appearance_mode_event(new_appearance_mode: str):
                            ctk.set_appearance_mode(new_appearance_mode)
                            fa = open("appearance.txt", "w")
                            fa.write(new_appearance_mode)
                            fa.close()

                        appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:")
                        appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                        appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                                   command=change_appearance_mode_event)
                        appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                        def change_scaling_event(new_scaling: str):
                            new_scaling_float = int(new_scaling.replace("%", "")) / 100
                            ctk.set_widget_scaling(new_scaling_float)
                            fs = open('Scale.txt', 'w')
                            n_scaling = new_scaling.replace("%", "")
                            fs.write(n_scaling)
                            fs.close()

                        scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:")
                        scaling_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)
                        scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                           command=change_scaling_event)
                        scaling_option.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)

                        def destroy_settings():
                            settings_frame.destroy()
                            settings_1.destroy()
                            settings_label.destroy()
                            appearance_mode_label.destroy()
                            appearance_mode_option.destroy()
                            scaling_label.destroy()
                            scaling_option.destroy()
                            _save.destroy()

                        _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15),
                                              command=destroy_settings)
                        _save.pack(padx=10, pady=12)
                        _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

                    settings_button = ctk.CTkButton(master=teach_frame, text="Settings", font=('Segoe Ui', 15),
                                                    command=settings)
                    settings_button.pack(padx=10, pady=12)
                    settings_button.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

                    _quit = ctk.CTkButton(master=teach_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
                    _quit.pack(padx=10, pady=12)
                    _quit.place(relx=0.5, rely=0.7, anchor=ctk.CENTER)
                else:
                    teach_frame.destroy()

            teach2_checkbox = ctk.CTkCheckBox(master=frame, text='Teacher', corner_radius=5, font=('Segoe Ui', 21),
                                              variable=acb_va2, onvalue=1, offvalue=0, command=account2)
            teach2_checkbox.pack()
            teach2_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

            user2_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

            teach2_label = ctk.CTkLabel(master=frame, text="", image=user2_image)
            teach2_label.pack()
            teach2_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

            path = "timetable_teachers.xlsx"
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook['Class 11 B']
            header = sheet.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            body = sheet.iter_rows(min_row=2, max_row=6, max_col=9, values_only=True)
            header = [r for r in header]
            body = [r for r in body]
            workbook.close()

            style = ttk.Style()
            style.configure("Treeview", rowheight=60)
            style.configure("Treeview.Heading", font=('Segoe UI', 20))

            tree = ttk.Treeview(master=frame, selectmode='browse', show='headings', columns=header[0], height=5)
            tree.pack(padx=10, pady=10, fill='x')
            tree.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            tree.tag_configure("tag_item", font=('Segoe Ui', 12), foreground='black')

            for i in header[0]:
                tree.column(i, anchor=ctk.CENTER)
                tree.heading(i, text=i)

            for i in body:
                tree.insert('', 'end', iid=i[0], values=i, tags="tag_item")

            text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
            text_name.pack(padx=1, pady=2)
            text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

            def _back():
                view2_label.destroy()
                view2_back.destroy()
                text_name.destroy()
                class_label.destroy()
                tree.destroy()
                teach2_checkbox.destroy()
                teach2_label.destroy()
                if acb_va2.get() == 1:
                    teach_frame.destroy()
                class11()

            view2_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
            view2_back.pack(pady=60)
            view2_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

        def go_view1():
            class10_label.destroy()
            class10_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class10_name.destroy()
            class10_back.destroy()
            teach1_checkbox.destroy()
            teach1_label.destroy()
            if acb_va1.get() == 1:
                teach_frame.destroy()
            view1()

        def go_view2():
            class10_label.destroy()
            class10_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class10_name.destroy()
            class10_back.destroy()
            teach1_label.destroy()
            teach1_checkbox.destroy()
            if acb_va1.get() == 1:
                teach_frame.destroy()
            view2()

        view1_button = ctk.CTkButton(frame, text="Class 10 A", font=('Segoe Ui', 25), command=go_view1)
        view1_button.pack(padx=10, pady=30)
        view1_button.place(relx=0.5, rely=0.45, anchor=ctk.N)

        view2_button = ctk.CTkButton(frame, text="Class 10 C", font=('Segoe Ui', 25), command=go_view2)
        view2_button.pack(padx=10, pady=30)
        view2_button.place(relx=0.5, rely=0.55, anchor=ctk.N)

        def back10():
            class10_label.destroy()
            class10_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class10_name.destroy()
            class10_back.destroy()
            teach1_checkbox.destroy()
            teach1_label.destroy()
            if acb_va1.get() == 1:
                teach_frame.destroy()
            tabview()

        class10_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
        class10_name.pack(padx=1, pady=2)
        class10_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

        class10_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=back10)
        class10_back.pack(pady=60)
        class10_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

    def go_class10():
        view_label.destroy()
        view_label2.destroy()
        view_back.destroy()
        class10_button.destroy()
        class11_button.destroy()
        teach_checkbox.destroy()
        teach_label.destroy()
        _name.destroy()
        if acb_va.get() == 1:
            teach_frame.destroy()
        class10()

    def go_class11():
        view_label.destroy()
        view_label2.destroy()
        view_back.destroy()
        class10_button.destroy()
        class11_button.destroy()
        teach_label.destroy()
        teach_checkbox.destroy()
        _name.destroy()
        if acb_va.get() == 1:
            teach_frame.destroy()
        class11()

    class10_button = ctk.CTkButton(frame, text="Class 10", font=('Segoe Ui', 25), command=go_class10)
    class10_button.pack(padx=10, pady=30)
    class10_button.place(relx=0.5, rely=0.45, anchor=ctk.N)

    class11_button = ctk.CTkButton(frame, text="Class 11", font=('Segoe Ui', 25), command=go_class11)
    class11_button.pack(padx=10, pady=30)
    class11_button.place(relx=0.5, rely=0.55, anchor=ctk.N)

    _name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
    _name.pack(padx=1, pady=2)
    _name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

    def back():
        view_label.destroy()
        view_label2.destroy()
        view_back.destroy()
        class10_button.destroy()
        class11_button.destroy()
        teach_label.destroy()
        teach_checkbox.destroy()
        _name.destroy()
        if acb_va.get() == 1:
            teach_frame.destroy()
        tab2()

    view_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=back)
    view_back.pack(pady=60)
    view_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)


# to view the present day timetable - Teachers
def ttabsub():
    sub_label = ctk.CTkLabel(master=frame, text="Today's Timetables", font=('Segoe Ui', 40))
    sub_label.pack(padx=20, pady=100)
    sub_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

    acb_va = ctk.IntVar(value=0)

    # noinspection PyUnboundLocalVariable,PyGlobalUndefined
    def account():
        if acb_va.get() == 1:

            global teach_frame

            teach_frame = ctk.CTkFrame(master=frame, width=150, height=100, corner_radius=15)
            teach_frame.pack()
            teach_frame.place(relx=0.895, rely=0.08, anchor=ctk.NW)

            def settings():
                settings_frame = ctk.CTkFrame(master=frame, corner_radius=15)
                settings_frame.pack(fill="both", expand=True)

                settings_1 = ctk.CTkFrame(master=settings_frame, corner_radius=30)
                settings_1.pack(pady=60, padx=550, fill="both", expand=True)

                settings_label = ctk.CTkLabel(master=settings_1, text="Settings", font=('Segoe Ui', 55))
                settings_label.pack(padx=10, pady=12)
                settings_label.place(relx=0.5, rely=0.08, anchor=ctk.N)

                def change_appearance_mode_event(new_appearance_mode: str):
                    ctk.set_appearance_mode(new_appearance_mode)
                    fa = open("appearance.txt", "w")
                    fa.write(new_appearance_mode)
                    fa.close()

                appearance_mode_label = ctk.CTkLabel(master=settings_1, text="Appearance Mode:", font=('Segoe Ui', 20))
                appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
                appearance_mode_option = ctk.CTkOptionMenu(settings_1, values=["Dark", "Light", "System"],
                                                           command=change_appearance_mode_event, font=('Segoe Ui', 15))
                appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

                def change_scaling_event(new_scaling: str):
                    new_scaling_float = int(new_scaling.replace("%", "")) / 100
                    ctk.set_widget_scaling(new_scaling_float)
                    fs = open('Scale.txt', 'w')
                    n_scaling = new_scaling.replace("%", "")
                    fs.write(n_scaling)
                    fs.close()

                scaling_label = ctk.CTkLabel(master=settings_1, text="UI Scaling:", font=('Segoe Ui', 20))
                scaling_label.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)
                scaling_option = ctk.CTkOptionMenu(master=settings_1, values=["100%", "110%", "120%"],
                                                   command=change_scaling_event, font=('Segoe Ui', 15))
                scaling_option.place(relx=0.5, rely=0.6, anchor=ctk.CENTER)

                def destroy_settings():
                    settings_frame.destroy()
                    settings_1.destroy()
                    settings_label.destroy()
                    appearance_mode_label.destroy()
                    appearance_mode_option.destroy()
                    scaling_label.destroy()
                    scaling_option.destroy()
                    _save.destroy()

                _save = ctk.CTkButton(master=settings_1, text="Save", font=('Segoe Ui', 15), command=destroy_settings)
                _save.pack(padx=10, pady=12)
                _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

            settings_button = ctk.CTkButton(master=teach_frame, text="Settings", font=('Segoe Ui', 15),
                                            command=settings)
            settings_button.pack(padx=10, pady=12)
            settings_button.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

            _quit = ctk.CTkButton(master=teach_frame, text="Quit", font=('Segoe Ui', 15), command=exit)
            _quit.pack(padx=10, pady=12)
            _quit.place(relx=0.5, rely=0.7, anchor=ctk.CENTER)
        else:
            teach_frame.destroy()

    teach_checkbox = ctk.CTkCheckBox(master=frame, text='Teacher', corner_radius=5, font=('Segoe Ui', 21),
                                     variable=acb_va, onvalue=1, offvalue=0, command=account)
    teach_checkbox.pack()
    teach_checkbox.place(relx=0.915, rely=0.031, anchor=ctk.NW)

    user_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

    teach_label = ctk.CTkLabel(master=frame, text="", image=user_image)
    teach_label.pack()
    teach_label.place(relx=0.91, rely=0.03, anchor=ctk.NW)

    wip_label = ctk.CTkLabel(master=frame, text="Work in progress", font=('Segoe Ui', 35))
    wip_label.pack(padx=20, pady=100)
    wip_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

    wip_label2 = ctk.CTkLabel(master=frame, text="will be added later on...", font=('Segoe Ui', 20))
    wip_label2.pack(padx=20, pady=10)
    wip_label2.place(relx=0.5, rely=0.58, anchor=ctk.CENTER)

    text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
    text_name.pack(padx=1, pady=2)
    text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

    def _back():
        sub_label.destroy()
        sub_back.destroy()
        wip_label.destroy()
        wip_label2.destroy()
        text_name.destroy()
        ttab2()

    sub_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
    sub_back.pack(pady=60)
    sub_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)


# to check login
def login():
    fp = open('login_remember_admin.txt', 'r')
    rl = fp.read()
    fp2 = open('login remember teach.txt', 'r')
    rl2 = fp2.read()
    if rl == "true":
        tab2()
    elif rl2 == "true":
        ttab2()
    else:
        tab1()


# starting of the program
def starting():
    heading = ctk.CTkLabel(master=frame, text="Time Management System", font=('Segoe Ui', 45))
    heading.pack(padx=10, pady=10)
    heading.place(relx=0.5, rely=0.08, anchor=ctk.N)

    heading2 = ctk.CTkLabel(master=frame, text="Helps you manage your timetables :)", font=('Segoe Ui', 20))
    heading2.pack(padx=10, pady=10)
    heading2.place(relx=0.5, rely=0.17, anchor=ctk.N)

    members = ctk.CTkLabel(master=frame, text=name, font=('Segoe Ui', 30))
    members.pack(padx=10, pady=10)
    members.place(relx=0.5, rely=0.86, anchor=ctk.S)

    team = "Rohan and Libin"
    members2 = ctk.CTkLabel(master=frame, text=team, font=('Segoe Ui', 17))
    members2.pack(padx=10, pady=10)
    members2.place(relx=0.5, rely=0.9, anchor=ctk.S)

    def enter_program():
        heading.destroy()
        heading2.destroy()
        members.destroy()
        members2.destroy()
        ebutton.destroy()
        _quit.destroy()
        login()

    ebutton = ctk.CTkButton(frame, text="Enter", font=('Segoe Ui', 22), command=enter_program)
    ebutton.pack(padx=10, pady=10)
    ebutton.place(relx=0.5, rely=0.5, anchor=ctk.S)

    _quit = ctk.CTkButton(frame, text="Quit", font=('Segoe Ui', 22), command=exit)
    _quit.pack(padx=10, pady=10)
    _quit.place(relx=0.5, rely=0.52, anchor=ctk.N)


starting()

root.mainloop()
