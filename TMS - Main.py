"""The Time Management System - C.S Project"""

# Importing required modules
import customtkinter as ctk
import tkinter
from tkinter import ttk
from PIL import ImageTk
from PIL import Image
import openpyxl

# importing the setting of the program
fam = open("appearance.txt", "r")
appearance = fam.read()
fam.close()

fsm = open('Scale.txt', 'r')
scaling = fsm.read()
fsm.close()
file_scaling = int(scaling) / 100

# setting the theme of the program
ctk.set_appearance_mode(appearance)
ctk.set_default_color_theme("green")
ctk.set_widget_scaling(file_scaling)

# making the program fullscreen only
root = ctk.CTk()
root.title("TMS")
# root.overrideredirect(True)
# root.attributes('-fullscreen', True)
root.geometry("{0}x{1}+0+0".format(root.winfo_screenwidth(), root.winfo_screenheight()))

# setting the frame of the program
frame = ctk.CTkFrame(master=root)

def create_frame():

    global frame

    frame = ctk.CTkFrame(master=root)
    frame.pack(fill="both", expand=True)

create_frame()

# dev team name x_x
name = "The Mallus"

# defined y location for the login form to make adjustment easier
login_form_y = 0.45

label1 = ctk.CTkLabel(master=frame, text="", font=('Segoe Ui', 45))

# user profile for admin
def account():
    global frame
    is_dropdown_open = False

    user_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

    acc_frame_margin = 16
    acc_frame_padding = 8

    acc_frame = ctk.CTkFrame(master=frame, height=45, corner_radius=10, fg_color="#383838")
    acc_frame.pack(side=ctk.TOP, anchor=ctk.E, padx=acc_frame_margin, pady=acc_frame_margin)

    dropdown_frame = ctk.CTkFrame(master=frame, corner_radius=15, fg_color="#383838")

    def open_settings():
        settings_frame = ctk.CTkFrame(master=frame, fg_color="#333333")
        settings_frame.place(relx=0, rely=0, relheight=1, relwidth=1)

        settings_label = ctk.CTkLabel(master=settings_frame, text="Settings", font=('Segoe Ui', 45))
        settings_label.pack(padx=10, pady=(60, 0))

        settings_box_container = ctk.CTkFrame(master=settings_frame, fg_color="#333333")
        settings_box_container.grid_rowconfigure(0, weight=1)
        settings_box_container.grid_columnconfigure(0, weight=1)
        settings_box_container.pack(fill="both", expand=True)

        settings_box = ctk.CTkFrame(master=settings_box_container, corner_radius=30)
        settings_box.grid(row=0, column=0, sticky="", ipadx=64, ipady=16)

        # settings_box.pack(ipadx=64, ipady=16, anchor=ctk.CENTER)

        def change_appearance_mode_event(new_appearance_mode: str):
            ctk.set_appearance_mode(new_appearance_mode)
            fa = open("appearance.txt", "w")
            fa.write(new_appearance_mode)
            fa.close()

        appearance_mode_label = ctk.CTkLabel(master=settings_box, text="Appearance Mode:", font=('Segoe Ui', 20))
        appearance_mode_label.pack(padx=10, pady=(20, 10))
        # appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
        appearance_mode_option = ctk.CTkOptionMenu(settings_box, values=["Dark", "Light", "System"],
                                                   command=change_appearance_mode_event, font=('Segoe Ui', 15))
        appearance_mode_option.pack(padx=10, pady=10)

        # appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

        def change_scaling_event(new_scaling: str):
            new_scaling = new_scaling.replace("%", "")
            ctk.set_widget_scaling(int(new_scaling)/100)
            fs = open('Scale.txt', 'w')
            fs.write(new_scaling)
            fs.close()

        scaling_label = ctk.CTkLabel(master=settings_box, text="UI Scaling:", font=('Segoe Ui', 20))
        scaling_label.pack(padx=10, pady=10)
        # scaling_label.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)
        scaling_option = ctk.CTkOptionMenu(master=settings_box, values=["100%", "110%", "120%"],
                                           command=change_scaling_event, font=('Segoe Ui', 15))
        scaling_option.pack(padx=10, pady=10)

        # scaling_option.place(relx=0.5, rely=0.6, anchor=ctk.CENTER)

        def destroy_settings():
            settings_frame.destroy()
            settings_box.destroy()
            settings_label.destroy()
            appearance_mode_label.destroy()
            appearance_mode_option.destroy()
            scaling_label.destroy()
            scaling_option.destroy()
            _save.destroy()

        _save = ctk.CTkButton(master=settings_box, text="Save", font=('Segoe Ui', 15), command=destroy_settings)
        _save.pack(padx=10, pady=(20, 10))
        # _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

    def logout():
        fh = open('login_remember_admin.txt', 'w')
        str2 = "false"
        print(str2)
        fh.write(str2)
        fh.close()
        label1.pack_forget()
        frame.destroy()
        create_frame()
        tab1()
        destroy_dropdown(True)

    settings_button = ctk.CTkButton(master=dropdown_frame, text="Settings", font=('Segoe Ui', 15),
                                    command=open_settings)

    logout1 = ctk.CTkButton(master=dropdown_frame, text="Logout", font=('Segoe Ui', 15), command=logout)

    _quit = ctk.CTkButton(master=dropdown_frame, text="Quit", font=('Segoe Ui', 15), command=exit)

    def place_dropdown():
        dropdown_frame.pack(side=ctk.TOP, anchor=ctk.NE, padx=acc_frame_margin)

        settings_button.pack(padx=10, pady=10)
        logout1.pack(padx=10)
        _quit.pack(padx=10, pady=10)

    def destroy_dropdown(do_logout=False):
        dropdown_frame.pack_forget()
        if do_logout:
            acc_frame.pack_forget()

    def toggle_dropdown():
        nonlocal is_dropdown_open
        destroy_dropdown() if is_dropdown_open else place_dropdown()
        is_dropdown_open = not is_dropdown_open

    # user_icon = ctk.CTkLabel(master=acc_frame, text="", image=user_image)
    # user_icon.pack(side=ctk.LEFT, padx=acc_frame_padding)

    user_button = ctk.CTkButton(master=acc_frame, image=user_image, text="  Admin", font=('Segoe Ui', 18),
                                fg_color="#383838", hover_color="#383838",
                                command=toggle_dropdown, anchor="w", compound="left")
    user_button.pack(padx=acc_frame_padding, pady=acc_frame_padding)

# user profile for teachers
def taccount():
    global frame
    is_dropdown_open = False

    user_image = ctk.CTkImage(Image.open("User.png"), size=(32, 32))

    acc_frame_margin = 16
    acc_frame_padding = 8

    acc_frame = ctk.CTkFrame(master=frame, height=45, corner_radius=10, fg_color="#383838")
    acc_frame.pack(side=ctk.TOP, anchor=ctk.E, padx=acc_frame_margin, pady=acc_frame_margin)

    dropdown_frame = ctk.CTkFrame(master=frame, corner_radius=15, fg_color="#383838")

    def open_settings():
        settings_frame = ctk.CTkFrame(master=frame, fg_color="#333333")
        settings_frame.place(relx=0, rely=0, relheight=1, relwidth=1)

        settings_label = ctk.CTkLabel(master=settings_frame, text="Settings", font=('Segoe Ui', 45))
        settings_label.pack(padx=10, pady=(60, 0))

        settings_box_container = ctk.CTkFrame(master=settings_frame, fg_color="#333333")
        settings_box_container.grid_rowconfigure(0, weight=1)
        settings_box_container.grid_columnconfigure(0, weight=1)
        settings_box_container.pack(fill="both", expand=True)

        settings_box = ctk.CTkFrame(master=settings_box_container, corner_radius=30)
        settings_box.grid(row=0, column=0, sticky="", ipadx=64, ipady=16)

        # settings_box.pack(ipadx=64, ipady=16, anchor=ctk.CENTER)

        def change_appearance_mode_event(new_appearance_mode: str):
            ctk.set_appearance_mode(new_appearance_mode)
            fa = open("appearance.txt", "w")
            fa.write(new_appearance_mode)
            fa.close()

        appearance_mode_label = ctk.CTkLabel(master=settings_box, text="Appearance Mode:", font=('Segoe Ui', 20))
        appearance_mode_label.pack(padx=10, pady=(20, 10))
        # appearance_mode_label.place(relx=0.5, rely=0.4, anchor=ctk.CENTER)
        appearance_mode_option = ctk.CTkOptionMenu(settings_box, values=["Dark", "Light", "System"],
                                                   command=change_appearance_mode_event, font=('Segoe Ui', 15))
        appearance_mode_option.pack(padx=10, pady=10)

        # appearance_mode_option.place(relx=0.5, rely=0.45, anchor=ctk.CENTER)

        def change_scaling_event(new_scaling: str):
            new_scaling_float = int(new_scaling.replace("%", "")) / 100
            ctk.set_widget_scaling(new_scaling_float)
            fs = open('Scale.txt', 'w')
            n_scaling = new_scaling.replace("%", "")
            fs.write(n_scaling)
            fs.close()

        scaling_label = ctk.CTkLabel(master=settings_box, text="UI Scaling:", font=('Segoe Ui', 20))
        scaling_label.pack(padx=10, pady=10)
        # scaling_label.place(relx=0.5, rely=0.55, anchor=ctk.CENTER)
        scaling_option = ctk.CTkOptionMenu(master=settings_box, values=["100%", "110%", "120%"],
                                           command=change_scaling_event, font=('Segoe Ui', 15))
        scaling_option.pack(padx=10, pady=10)

        # scaling_option.place(relx=0.5, rely=0.6, anchor=ctk.CENTER)

        def destroy_settings():
            settings_frame.destroy()
            settings_box.destroy()
            settings_label.destroy()
            appearance_mode_label.destroy()
            appearance_mode_option.destroy()
            scaling_label.destroy()
            scaling_option.destroy()
            _save.destroy()

        _save = ctk.CTkButton(master=settings_box, text="Save", font=('Segoe Ui', 15), command=destroy_settings)
        _save.pack(padx=10, pady=(20, 10))
        # _save.place(relx=0.5, rely=0.8, anchor=ctk.CENTER)

    def logout():
        fh = open('login_remember_teach.txt', 'w')
        str2 = "false"
        print(str2)
        fh.write(str2)
        fh.close()
        label1.pack_forget()
        frame.destroy()
        create_frame()
        tab1()
        destroy_dropdown(True)

    settings_button = ctk.CTkButton(master=dropdown_frame, text="Settings", font=('Segoe Ui', 15),
                                    command=open_settings)

    logout1 = ctk.CTkButton(master=dropdown_frame, text="Logout", font=('Segoe Ui', 15), command=logout)

    _quit = ctk.CTkButton(master=dropdown_frame, text="Quit", font=('Segoe Ui', 15), command=exit)

    def place_dropdown():
        dropdown_frame.pack(side=ctk.TOP, anchor=ctk.NE, padx=acc_frame_margin)

        settings_button.pack(padx=10, pady=10)
        logout1.pack(padx=10)
        _quit.pack(padx=10, pady=10)

    def destroy_dropdown(do_logout=False):
        dropdown_frame.pack_forget()
        if do_logout:
            acc_frame.pack_forget()

    def toggle_dropdown():
        nonlocal is_dropdown_open
        destroy_dropdown() if is_dropdown_open else place_dropdown()
        is_dropdown_open = not is_dropdown_open

    # user_icon = ctk.CTkLabel(master=acc_frame, text="", image=user_image)
    # user_icon.pack(side=ctk.LEFT, padx=acc_frame_padding)

    user_button = ctk.CTkButton(master=acc_frame, image=user_image, text="  Teacher", font=('Segoe Ui', 18),
                                fg_color="#383838", hover_color="#383838",
                                command=toggle_dropdown, anchor="w", compound="left")
    user_button.pack(padx=acc_frame_padding, pady=acc_frame_padding)

# login program
# if rl is false is runs the login program
def tab1():

    frame.pack_configure(fill="both", expand=True)

    label = ctk.CTkLabel(master=frame, text="Login Page", font=('Segoe Ui', 45))
    label.pack(fill="both", expand=True)
    label.place(relx=0.5, rely=0.08, anchor=ctk.N)

    frame1 = ctk.CTkFrame(master=frame, width=320, height=360, corner_radius=20)
    frame1.place(relx=0.5, rely=0.5, anchor=tkinter.CENTER)

    label1 = ctk.CTkLabel(master=frame1, text="Log into your Account", font=('Segoe Ui', 25))
    label1.pack()
    label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

    entry1 = ctk.CTkEntry(master=frame1, width=250, placeholder_text="Username", font=('Segoe Ui', 17))
    entry1.pack(padx=10, pady=12)
    entry1.place(relx=0.5, rely=login_form_y - 0.08, anchor=ctk.CENTER)

    entry2 = ctk.CTkEntry(master=frame1, width=250, placeholder_text="Password", show="*", font=('Segoe Ui', 17))
    entry2.pack(padx=10, pady=12)
    entry2.place(relx=0.5, rely=login_form_y + 0.07, anchor=ctk.CENTER)

    cb_va = ctk.IntVar(value=0)

    def my_show():
        if cb_va.get() == 1:
            entry2.configure(show='')
        else:
            entry2.configure(show='*')

    checkbox1 = ctk.CTkCheckBox(master=frame1, text='', corner_radius=5, font=('Segoe Ui', 12), variable=cb_va, onvalue=1, offvalue=0, command=my_show)
    checkbox1.configure(width=0, height=1)
    checkbox1.place(relx=0.836, rely=login_form_y + 0.07, anchor=ctk.CENTER)

    def p_forgot():
        textbox4
        textbox4.place(relx=0.5, rely=login_form_y + 0.3, anchor=ctk.CENTER)
        textbox3.place(relx=2, rely=2)

    forgot = ctk.CTkButton(master=frame1, text="Forgot Password?", font=('Century Gothic', 12), fg_color="#333333", command=p_forgot)
    forgot.configure(height=1, width=1)
    forgot.place(relx=0.75, rely=login_form_y + 0.2, anchor=ctk.CENTER)

    check_state = ctk.IntVar()

    def check_login():
        if entry1.get() == "Admin" and entry2.get() == "Admin123":
            if check_state.get() == 1:
                fh = open('login_remember_admin.txt', 'w')
                str1 = "true"
                print(str1)
                fh.write(str1)
                fh.close()
            elif check_state.get() == 0:
                fh = open('login_remember_admin.txt', 'w')
                str2 = "false"
                print(str2)
                fh.write(str2)
                fh.close()
        if entry1.get() == "Teacher" and entry2.get() == "Teacher123":
            if check_state.get() == 1:
                fh = open('login_remember_teach.txt', 'w')
                str1 = "true"
                print(str1)
                fh.write(str1)
                fh.close()
            elif check_state.get() == 0:
                fh = open('login_remember_teach.txt', 'w')
                str2 = "false"
                print(str2)
                fh.write(str2)
                fh.close()

    checkbox2 = ctk.CTkCheckBox(master=frame1, text="Remember Me", font=('Century Gothic', 12), variable=check_state, command=check_login)
    checkbox2.pack(padx=10, pady=12)
    checkbox2.configure(height=0, width=0)
    checkbox2.place(relx=0.3, rely=login_form_y + 0.2, anchor=ctk.CENTER)

    textbox3 = ctk.CTkLabel(master=frame1, height=5, text="*Incorrect Username or Password*", font=('Segoe Ui', 13))

    textbox4 = ctk.CTkLabel(master=frame1, height=5, text="How did you forget that? :|", font=('Segoe Ui', 13))

    def open_admin():
        entry2.configure(show='*')
        label.destroy()
        label1.destroy()
        frame1.configure(width=0, height=0)
        entry1.destroy()
        entry2.destroy()
        checkbox1.destroy()
        checkbox2.destroy()
        textbox3.destroy()
        textbox4.destroy()
        forgot.destroy()
        cbutton.destroy()
        _quit.destroy()
        text_name.destroy()
        tab2()
        account()

    def open_teacher():
        entry2.configure(show='*')
        label.destroy()
        label1.destroy()
        frame1.configure(width=0, height=0)
        entry1.destroy()
        entry2.destroy()
        checkbox1.destroy()
        checkbox2.destroy()
        textbox3.destroy()
        textbox4.destroy()
        forgot.destroy()
        cbutton.destroy()
        _quit.destroy()
        text_name.destroy()
        ttab2()
        taccount()

    def login_details():
        if entry1.get() == "Admin" and entry2.get() == "Admin123":
            open_admin()
        elif entry1.get() == "Teacher" and entry2.get() == "Teacher123":
            open_teacher()
        else:
            textbox3
            textbox3.place(relx=0.5, rely=login_form_y + 0.3, anchor=ctk.CENTER)
            textbox4.place(relx=2, rely=2)

    cbutton = ctk.CTkButton(master=frame1, width=180, height=15, corner_radius=100, text="Login", font=('Segoe Ui', 20), command=login_details)
    cbutton.pack(padx=10, pady=12)
    cbutton.place(relx=0.5, rely=login_form_y + 0.43, anchor=ctk.CENTER)

    _quit = ctk.CTkButton(master=frame, text="Quit", font=('Segoe Ui', 15), command=exit)
    _quit.pack()
    _quit.place(relx=0.89, rely=0.95, anchor=ctk.NW)

    text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
    text_name.pack(padx=1, pady=2)
    text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

# program lobby - Admin
# if rl is true it skips login and runs the program
def tab2():

    frame.pack_configure(fill="both", expand=True)

    global label1

    label1.configure(text="Time Management System")
    label1.pack(fill="both", expand=True)
    label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

    def clear_tab2():
        view_button.destroy()
        man_button.destroy()
        sub_button.destroy()
        text_name.destroy()

    def button_view():
        clear_tab2()
        tabview()

    def button_man():
        clear_tab2()
        tabman()

    def button_sub():
        clear_tab2()
        tabsub()

    view_button = ctk.CTkButton(frame, text="View Timetables", font=('Segoe Ui', 35), command=button_view)
    view_button.pack(padx=10, pady=30)
    view_button.place(relx=0.5, rely=0.35, anchor=ctk.N)

    man_button = ctk.CTkButton(frame, text="Manage Timetables", font=('Segoe Ui', 35), command=button_man)
    man_button.pack(padx=10, pady=30)
    man_button.place(relx=0.5, rely=0.5, anchor=ctk.N)

    sub_button = ctk.CTkButton(frame, text="Substitute Absent Teachers", font=('Segoe Ui', 35), command=button_sub)
    sub_button.pack(padx=10, pady=30)
    sub_button.place(relx=0.5, rely=0.65, anchor=ctk.N)

    text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
    text_name.pack(padx=1, pady=2)
    text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

# to view the timetable - Admin
def tabview():

    frame.pack_configure(fill="both", expand=True)

    label1.configure(text="Timetables")
    label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

    view_label2 = ctk.CTkLabel(master=frame, text="Pick a class", font=('Segoe Ui', 35))
    view_label2.pack()
    view_label2.place(relx=0.5, rely=0.2, anchor=ctk.N)

# reading the Excel files

    def class11():
        label1.configure(text="Timetables")
        label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

        class11_label2 = ctk.CTkLabel(master=frame, text="Class 11", font=('Segoe Ui', 35))
        class11_label2.place(relx=0.5, rely=0.2, anchor=ctk.N)

        def view1():

            label1.configure(text="Timetables")
            label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

            class_label = ctk.CTkLabel(master=frame, text="Class 11 A", font=('Segoe Ui', 35))
            class_label.place(relx=0.5, rely=0.2, anchor=ctk.N)

            path = "timetable_teachers.xlsx"
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook['Class 11 A']
            header = sheet.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            body = sheet.iter_rows(min_row=2, max_row=6, max_col=9, values_only=True)
            header = [r for r in header]
            body = [r for r in body]
            workbook.close()

            style = ttk.Style()
            style.configure("Treeview", rowheight=60)
            style.configure("Treeview.Heading", font=('Segoe UI', 20))

            tree = ttk.Treeview(master=frame, selectmode='browse', show='headings', columns=header[0], height=5)
            tree.pack(padx=10, pady=10, fill='x')
            tree.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            tree.tag_configure("tag_item", font=('Segoe Ui', 12), foreground='black')

            for i in header[0]:
                tree.column(i, anchor=ctk.CENTER)
                tree.heading(i, text=i)

            for i in body:
                tree.insert('', 'end', iid=i[0], values=i, tags="tag_item")

            text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
            text_name.pack(padx=1, pady=2)
            text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

            def _back():
                label1.pack_forget()
                view1_back.destroy()
                text_name.destroy()
                class_label.destroy()
                tree.destroy()
                class11()

            view1_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
            view1_back.pack(pady=60)
            view1_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

        def view2():

            label1.configure(text="Timetables")
            label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

            class_label = ctk.CTkLabel(master=frame, text="Class 11 C", font=('Segoe Ui', 35))
            class_label.pack()
            class_label.place(relx=0.5, rely=0.2, anchor=ctk.N)

            path = "timetable_teachers.xlsx"
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook['Class 11 B']
            header = sheet.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            body = sheet.iter_rows(min_row=2, max_row=6, max_col=9, values_only=True)
            header = [r for r in header]
            body = [r for r in body]
            workbook.close()

            style = ttk.Style()
            style.configure("Treeview", rowheight=60)
            style.configure("Treeview.Heading", font=('Segoe UI', 20))

            tree = ttk.Treeview(master=frame, selectmode='browse', show='headings', columns=header[0], height=5)
            tree.pack(padx=10, pady=10, fill='x')
            tree.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            tree.tag_configure("tag_item", font=('Segoe Ui', 12), foreground='black')

            for i in header[0]:
                tree.column(i, anchor=ctk.CENTER)
                tree.heading(i, text=i)

            for i in body:
                tree.insert('', 'end', iid=i[0], values=i, tags="tag_item")

            text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
            text_name.pack(padx=1, pady=2)
            text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

            def _back():
                label1.pack_forget()
                view2_back.destroy()
                text_name.destroy()
                class_label.destroy()
                tree.destroy()
                class11()

            view2_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
            view2_back.pack(pady=60)
            view2_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

        def go_view1():
            label1.pack_forget()
            class11_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class11_name.destroy()
            class11_back.destroy()
            view1()

        def go_view2():
            label1.pack_forget()
            class11_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class11_name.destroy()
            class11_back.destroy()
            view2()

        view1_button = ctk.CTkButton(frame, text="Class 11 A", font=('Segoe Ui', 25), command=go_view1)
        view1_button.pack(padx=10, pady=30)
        view1_button.place(relx=0.5, rely=0.45, anchor=ctk.N)

        view2_button = ctk.CTkButton(frame, text="Class 11 C", font=('Segoe Ui', 25), command=go_view2)
        view2_button.pack(padx=10, pady=30)
        view2_button.place(relx=0.5, rely=0.55, anchor=ctk.N)

        def back11():
            label1.pack_forget()
            class11_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class11_name.destroy()
            class11_back.destroy()
            tabview()

        class11_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
        class11_name.pack(padx=1, pady=2)
        class11_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

        class11_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=back11)
        class11_back.pack()
        class11_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

    def class12():
        label1.configure(text="Timetables")
        label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

        class12_label2 = ctk.CTkLabel(master=frame, text="Class 12", font=('Segoe Ui', 35))
        class12_label2.pack()
        class12_label2.place(relx=0.5, rely=0.2, anchor=ctk.N)

        def view1():

            label1.configure(text="Timetables")
            label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

            class_label = ctk.CTkLabel(master=frame, text="Class 12 A", font=('Segoe Ui', 35))
            class_label.pack()
            class_label.place(relx=0.5, rely=0.2, anchor=ctk.N)

            path = "timetable_teachers.xlsx"
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook['Class 11 A']
            header = sheet.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            body = sheet.iter_rows(min_row=2, max_row=6, max_col=9, values_only=True)
            header = [r for r in header]
            body = [r for r in body]
            workbook.close()

            style = ttk.Style()
            style.configure("Treeview", rowheight=60)
            style.configure("Treeview.Heading", font=('Segoe UI', 20))

            tree = ttk.Treeview(master=frame, selectmode='browse', show='headings', columns=header[0], height=5)
            tree.pack(padx=10, pady=10, fill='x')
            tree.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            tree.tag_configure("tag_item", font=('Segoe Ui', 12), foreground='black')

            for i in header[0]:
                tree.column(i, anchor=ctk.CENTER)
                tree.heading(i, text=i)

            for i in body:
                tree.insert('', 'end', iid=i[0], values=i, tags="tag_item")

            text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
            text_name.pack(padx=1, pady=2)
            text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

            def _back():
                label1.pack_forget()
                view1_back.destroy()
                text_name.destroy()
                class_label.destroy()
                tree.destroy()
                class12()

            view1_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
            view1_back.pack(pady=60)
            view1_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

        def view2():
            
            label1.configure(text="Timetables")
            label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

            class_label = ctk.CTkLabel(master=frame, text="Class 12 C", font=('Segoe Ui', 35))
            class_label.pack()
            class_label.place(relx=0.5, rely=0.2, anchor=ctk.N)

            path = "timetable_teachers.xlsx"
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook['Class 11 B']
            header = sheet.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            body = sheet.iter_rows(min_row=2, max_row=6, max_col=9, values_only=True)
            header = [r for r in header]
            body = [r for r in body]
            workbook.close()

            style = ttk.Style()
            style.configure("Treeview", rowheight=60)
            style.configure("Treeview.Heading", font=('Segoe UI', 20))

            tree = ttk.Treeview(master=frame, selectmode='browse', show='headings', columns=header[0], height=5)
            tree.pack(padx=10, pady=10, fill='x')
            tree.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            tree.tag_configure("tag_item", font=('Segoe Ui', 12), foreground='black')

            for i in header[0]:
                tree.column(i, anchor=ctk.CENTER)
                tree.heading(i, text=i)

            for i in body:
                tree.insert('', 'end', iid=i[0], values=i, tags="tag_item")

            text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
            text_name.pack(padx=1, pady=2)
            text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

            def _back():
                label1.pack_forget()
                view2_back.destroy()
                text_name.destroy()
                class_label.destroy()
                tree.destroy()
                class12()

            view2_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
            view2_back.pack(pady=60)
            view2_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

        def go_view1():
            label1.pack_forget()
            class12_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class12_name.destroy()
            class12_back.destroy()
            view1()

        def go_view2():
            label1.pack_forget()
            class12_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class12_name.destroy()
            class12_back.destroy()
            view2()

        view1_button = ctk.CTkButton(frame, text="Class 12 A", font=('Segoe Ui', 25), command=go_view1)
        view1_button.pack(padx=10, pady=30)
        view1_button.place(relx=0.5, rely=0.45, anchor=ctk.N)

        view2_button = ctk.CTkButton(frame, text="Class 12 C", font=('Segoe Ui', 25), command=go_view2)
        view2_button.pack(padx=10, pady=30)
        view2_button.place(relx=0.5, rely=0.55, anchor=ctk.N)

        def back12():
            label1.pack_forget()
            class12_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class12_name.destroy()
            class12_back.destroy()
            tabview()

        class12_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
        class12_name.pack(padx=1, pady=2)
        class12_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

        class12_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=back12)
        class12_back.pack(pady=60)
        class12_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

    def go_class11():
        label1.pack_forget()
        view_label2.destroy()
        view_back.destroy()
        class11_button.destroy()
        class12_button.destroy()
        class11()

    def go_class12():
        label1.pack_forget()
        view_label2.destroy()
        view_back.destroy()
        class11_button.destroy()
        class12_button.destroy()
        class12()

    class12_button = ctk.CTkButton(frame, text="Class 12", font=('Segoe Ui', 25), command=go_class12)
    class12_button.pack(padx=10, pady=30)
    class12_button.place(relx=0.5, rely=0.45, anchor=ctk.N)

    class11_button = ctk.CTkButton(frame, text="Class 11", font=('Segoe Ui', 25), command=go_class11)
    class11_button.pack(padx=10, pady=30)
    class11_button.place(relx=0.5, rely=0.55, anchor=ctk.N)

    _name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
    _name.pack(padx=1, pady=2)
    _name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

    def back():
        label1.pack_forget()
        view_label2.destroy()
        view_back.destroy()
        class12_button.destroy()
        class11_button.destroy()
        tab2()

    view_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=back)
    view_back.pack(pady=60)
    view_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

# to manage timetables - Admin
def tabman():

    frame.pack_configure(fill="both", expand=True)

    label1.configure(text="Manage Timetables")
    label1.pack(padx=20, pady=100)
    label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

    wip_label = ctk.CTkLabel(master=frame, text="Work in progress", font=('Segoe Ui', 35))
    wip_label.pack(padx=20, pady=100)
    wip_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

    wip_label2 = ctk.CTkLabel(master=frame, text="will be added later on...", font=('Segoe Ui', 20))
    wip_label2.pack(padx=20, pady=10)
    wip_label2.place(relx=0.5, rely=0.58, anchor=ctk.CENTER)

    text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
    text_name.pack(padx=1, pady=2)
    text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

    def _back():
        man_back.destroy()
        wip_label.destroy()
        wip_label2.destroy()
        text_name.destroy()
        tab2()

    man_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
    man_back.pack(pady=60)
    man_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

# to substitute absent teachers
def tabsub():

    frame.pack_configure(fill="both", expand=True)

    label1.configure(text="Substitute Absent Teachers")
    label1.pack(padx=20, pady=100)
    label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

    wip_label = ctk.CTkLabel(master=frame, text="Work in progress", font=('Segoe Ui', 35))
    wip_label.pack(padx=20, pady=100)
    wip_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

    text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
    text_name.pack(padx=1, pady=2)
    text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

    def _back():
        sub_back.destroy()
        wip_label.destroy()
        text_name.destroy()
        tab2()

    sub_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
    sub_back.pack(pady=60)
    sub_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

# program lobby - Teachers
# if username and password is entered for teachers it runs this code
def ttab2():

    frame.pack_configure(fill="both", expand=True)

    def destroy_ttab2():
        label1.pack_forget()
        view_button.destroy()
        sub_button.destroy()
        text_name.destroy()
        
    global label1

    label1.configure(text="Time Management System")
    label1.pack(fill="both", expand=True)
    label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

    def button_view():
        destroy_ttab2()
        ttabview()

    def button_sub():
        destroy_ttab2()
        ttabsub()

    view_button = ctk.CTkButton(frame, text="View Timetables", font=('Segoe Ui', 45), command=button_view)
    view_button.pack(padx=10, pady=30)
    view_button.place(relx=0.5, rely=0.43, anchor=ctk.N)

    sub_button = ctk.CTkButton(frame, text="View Today's Timetables", font=('Segoe Ui', 45), command=button_sub)
    sub_button.pack(padx=10, pady=30)
    sub_button.place(relx=0.5, rely=0.58, anchor=ctk.N)

    text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
    text_name.pack(padx=1, pady=2)
    text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

# to view the timetable - Teachers
def ttabview():

    frame.pack_configure(fill="both", expand=True)
    
    label1.configure(text="Timetables")
    label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

    view_label2 = ctk.CTkLabel(master=frame, text="Pick a class", font=('Segoe Ui', 35))
    view_label2.pack()
    view_label2.place(relx=0.5, rely=0.2, anchor=ctk.N)
    # reading the Excel files

    def class11():
        label1.configure(text="Timetables")
        label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

        class11_label2 = ctk.CTkLabel(master=frame, text="Class 11", font=('Segoe Ui', 35))
        class11_label2.place(relx=0.5, rely=0.2, anchor=ctk.N)

        def view1():

            label1.configure(text="Timetables")
            label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

            class_label = ctk.CTkLabel(master=frame, text="Class 11 A", font=('Segoe Ui', 35))
            class_label.place(relx=0.5, rely=0.2, anchor=ctk.N)

            path = "timetable_teachers.xlsx"
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook['Class 11 A']
            header = sheet.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            body = sheet.iter_rows(min_row=2, max_row=6, max_col=9, values_only=True)
            header = [r for r in header]
            body = [r for r in body]
            workbook.close()

            style = ttk.Style()
            style.configure("Treeview", rowheight=60)
            style.configure("Treeview.Heading", font=('Segoe UI', 20))

            tree = ttk.Treeview(master=frame, selectmode='browse', show='headings', columns=header[0], height=5)
            tree.pack(padx=10, pady=10, fill='x')
            tree.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            tree.tag_configure("tag_item", font=('Segoe Ui', 12), foreground='black')

            for i in header[0]:
                tree.column(i, anchor=ctk.CENTER)
                tree.heading(i, text=i)

            for i in body:
                tree.insert('', 'end', iid=i[0], values=i, tags="tag_item")

            text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
            text_name.pack(padx=1, pady=2)
            text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

            def _back():
                label1.pack_forget()
                view1_back.destroy()
                text_name.destroy()
                class_label.destroy()
                tree.destroy()
                class11()

            view1_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
            view1_back.pack(pady=60)
            view1_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

        def view2():

            label1.configure(text="Timetables")
            label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

            class_label = ctk.CTkLabel(master=frame, text="Class 11 C", font=('Segoe Ui', 35))
            class_label.pack()
            class_label.place(relx=0.5, rely=0.2, anchor=ctk.N)

            path = "timetable_teachers.xlsx"
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook['Class 11 B']
            header = sheet.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            body = sheet.iter_rows(min_row=2, max_row=6, max_col=9, values_only=True)
            header = [r for r in header]
            body = [r for r in body]
            workbook.close()

            style = ttk.Style()
            style.configure("Treeview", rowheight=60)
            style.configure("Treeview.Heading", font=('Segoe UI', 20))

            tree = ttk.Treeview(master=frame, selectmode='browse', show='headings', columns=header[0], height=5)
            tree.pack(padx=10, pady=10, fill='x')
            tree.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            tree.tag_configure("tag_item", font=('Segoe Ui', 12), foreground='black')

            for i in header[0]:
                tree.column(i, anchor=ctk.CENTER)
                tree.heading(i, text=i)

            for i in body:
                tree.insert('', 'end', iid=i[0], values=i, tags="tag_item")

            text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
            text_name.pack(padx=1, pady=2)
            text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

            def _back():
                label1.pack_forget()
                view2_back.destroy()
                text_name.destroy()
                class_label.destroy()
                tree.destroy()
                class11()

            view2_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
            view2_back.pack(pady=60)
            view2_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

        def go_view1():
            label1.pack_forget()
            class11_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class11_name.destroy()
            class11_back.destroy()
            view1()

        def go_view2():
            label1.pack_forget()
            class11_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class11_name.destroy()
            class11_back.destroy()
            view2()

        view1_button = ctk.CTkButton(frame, text="Class 11 A", font=('Segoe Ui', 25), command=go_view1)
        view1_button.pack(padx=10, pady=30)
        view1_button.place(relx=0.5, rely=0.45, anchor=ctk.N)

        view2_button = ctk.CTkButton(frame, text="Class 11 C", font=('Segoe Ui', 25), command=go_view2)
        view2_button.pack(padx=10, pady=30)
        view2_button.place(relx=0.5, rely=0.55, anchor=ctk.N)

        def back11():
            label1.pack_forget()
            class11_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class11_name.destroy()
            class11_back.destroy()
            ttabview()

        class11_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
        class11_name.pack(padx=1, pady=2)
        class11_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

        class11_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=back11)
        class11_back.pack()
        class11_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

    def class12():
        label1.configure(text="Timetables")
        label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

        class12_label2 = ctk.CTkLabel(master=frame, text="Class 12", font=('Segoe Ui', 35))
        class12_label2.pack()
        class12_label2.place(relx=0.5, rely=0.2, anchor=ctk.N)

        def view1():

            label1.configure(text="Timetables")
            label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

            class_label = ctk.CTkLabel(master=frame, text="Class 12 A", font=('Segoe Ui', 35))
            class_label.pack()
            class_label.place(relx=0.5, rely=0.2, anchor=ctk.N)

            path = "timetable_teachers.xlsx"
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook['Class 11 A']
            header = sheet.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            body = sheet.iter_rows(min_row=2, max_row=6, max_col=9, values_only=True)
            header = [r for r in header]
            body = [r for r in body]
            workbook.close()

            style = ttk.Style()
            style.configure("Treeview", rowheight=60)
            style.configure("Treeview.Heading", font=('Segoe UI', 20))

            tree = ttk.Treeview(master=frame, selectmode='browse', show='headings', columns=header[0], height=5)
            tree.pack(padx=10, pady=10, fill='x')
            tree.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            tree.tag_configure("tag_item", font=('Segoe Ui', 12), foreground='black')

            for i in header[0]:
                tree.column(i, anchor=ctk.CENTER)
                tree.heading(i, text=i)

            for i in body:
                tree.insert('', 'end', iid=i[0], values=i, tags="tag_item")

            text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
            text_name.pack(padx=1, pady=2)
            text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

            def _back():
                label1.pack_forget()
                view1_back.destroy()
                text_name.destroy()
                class_label.destroy()
                tree.destroy()
                class12()

            view1_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
            view1_back.pack(pady=60)
            view1_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

        def view2():
            
            label1.configure(text="Timetables")
            label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

            class_label = ctk.CTkLabel(master=frame, text="Class 12 C", font=('Segoe Ui', 35))
            class_label.pack()
            class_label.place(relx=0.5, rely=0.2, anchor=ctk.N)

            path = "timetable_teachers.xlsx"
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook['Class 11 B']
            header = sheet.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True)
            body = sheet.iter_rows(min_row=2, max_row=6, max_col=9, values_only=True)
            header = [r for r in header]
            body = [r for r in body]
            workbook.close()

            style = ttk.Style()
            style.configure("Treeview", rowheight=60)
            style.configure("Treeview.Heading", font=('Segoe UI', 20))

            tree = ttk.Treeview(master=frame, selectmode='browse', show='headings', columns=header[0], height=5)
            tree.pack(padx=10, pady=10, fill='x')
            tree.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

            tree.tag_configure("tag_item", font=('Segoe Ui', 12), foreground='black')

            for i in header[0]:
                tree.column(i, anchor=ctk.CENTER)
                tree.heading(i, text=i)

            for i in body:
                tree.insert('', 'end', iid=i[0], values=i, tags="tag_item")

            text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
            text_name.pack(padx=1, pady=2)
            text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

            def _back():
                label1.pack_forget()
                view2_back.destroy()
                text_name.destroy()
                class_label.destroy()
                tree.destroy()
                class12()

            view2_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
            view2_back.pack(pady=60)
            view2_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

        def go_view1():
            label1.pack_forget()
            class12_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class12_name.destroy()
            class12_back.destroy()
            view1()

        def go_view2():
            label1.pack_forget()
            class12_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class12_name.destroy()
            class12_back.destroy()
            view2()

        view1_button = ctk.CTkButton(frame, text="Class 12 A", font=('Segoe Ui', 25), command=go_view1)
        view1_button.pack(padx=10, pady=30)
        view1_button.place(relx=0.5, rely=0.45, anchor=ctk.N)

        view2_button = ctk.CTkButton(frame, text="Class 12 C", font=('Segoe Ui', 25), command=go_view2)
        view2_button.pack(padx=10, pady=30)
        view2_button.place(relx=0.5, rely=0.55, anchor=ctk.N)

        def back12():
            label1.pack_forget()
            class12_label2.destroy()
            view1_button.destroy()
            view2_button.destroy()
            class12_name.destroy()
            class12_back.destroy()
            ttabview()

        class12_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
        class12_name.pack(padx=1, pady=2)
        class12_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

        class12_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=back12)
        class12_back.pack(pady=60)
        class12_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

    def go_class11():
        label1.pack_forget()
        view_label2.destroy()
        view_back.destroy()
        class11_button.destroy()
        class12_button.destroy()
        class11()

    def go_class12():
        label1.pack_forget()
        view_label2.destroy()
        view_back.destroy()
        class11_button.destroy()
        class12_button.destroy()
        class12()
      
    class12_button = ctk.CTkButton(frame, text="Class 12", font=('Segoe Ui', 25), command=go_class12)
    class12_button.pack(padx=10, pady=30)
    class12_button.place(relx=0.5, rely=0.45, anchor=ctk.N)

    class11_button = ctk.CTkButton(frame, text="Class 11", font=('Segoe Ui', 25), command=go_class11)
    class11_button.pack(padx=10, pady=30)
    class11_button.place(relx=0.5, rely=0.55, anchor=ctk.N)

    _name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
    _name.pack(padx=1, pady=2)
    _name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

    def back():
        label1.pack_forget()
        view_label2.destroy()
        view_back.destroy()
        class12_button.destroy()
        class11_button.destroy()
        _name.destroy()
        ttab2()

    view_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=back)
    view_back.pack(pady=60)
    view_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

# to view the present day timetable - Teachers
def ttabsub():

    frame.pack_configure(fill="both", expand=True)

    label1.configure(text="Today's Timetables")
    label1.pack(padx=20, pady=100)
    label1.place(relx=0.5, rely=0.08, anchor=ctk.N)

    wip_label = ctk.CTkLabel(master=frame, text="Work in progress", font=('Segoe Ui', 35))
    wip_label.pack(padx=20, pady=100)
    wip_label.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

    wip_label2 = ctk.CTkLabel(master=frame, text="will be added later on...", font=('Segoe Ui', 20))
    wip_label2.pack(padx=20, pady=10)
    wip_label2.place(relx=0.5, rely=0.58, anchor=ctk.CENTER)

    text_name = ctk.CTkLabel(master=frame, height=5, text=name, font=('Segoe Ui', 18))
    text_name.pack(padx=1, pady=2)
    text_name.place(relx=0.065, rely=0.96, anchor=ctk.NE)

    def _back():
        label1.pack_forget()
        sub_back.destroy()
        wip_label.destroy()
        wip_label2.destroy()
        text_name.destroy()
        ttab2()

    sub_back = ctk.CTkButton(master=frame, text="Back", font=('Segoe Ui', 10), command=_back)
    sub_back.pack(pady=60)
    sub_back.place(relx=0.89, rely=0.95, anchor=ctk.NW)

# to check login
def login():

    fp = open('login_remember_admin.txt', 'r')
    rl = fp.read()
    fp2 = open('login_remember_teach.txt', 'r')
    rl2 = fp2.read()
    if rl == "true":
        tab2()
        account()
    elif rl2 == "true":
        ttab2()
        taccount()
    else:
        tab1()

# starting of the program
def starting():

    heading = ctk.CTkLabel(master=frame, text="Time Management System", font=('Segoe Ui', 45))
    heading.pack(padx=10, pady=10)
    heading.place(relx=0.5, rely=0.08, anchor=ctk.N)

    heading2 = ctk.CTkLabel(master=frame, text="Helps you manage your timetables :)", font=('Segoe Ui', 20))
    heading2.pack(padx=10, pady=10)
    heading2.place(relx=0.5, rely=0.17, anchor=ctk.N)

    members = ctk.CTkLabel(master=frame, text=name, font=('Segoe Ui', 30))
    members.pack(padx=10, pady=10)
    members.place(relx=0.5, rely=0.86, anchor=ctk.S)

    team = "Rohan and Libin"
    members2 = ctk.CTkLabel(master=frame, text=team, font=('Segoe Ui', 17))
    members2.pack(padx=10, pady=10)
    members2.place(relx=0.5, rely=0.9, anchor=ctk.S)

    def enter_program():
        heading.destroy()
        heading2.destroy()
        members.destroy()
        members2.destroy()
        ebutton.destroy()
        _quit.destroy()
        login()

    ebutton = ctk.CTkButton(frame, text="Enter", font=('Segoe Ui', 22), command=enter_program)
    ebutton.pack(padx=10, pady=10)
    ebutton.place(relx=0.5, rely=0.5, anchor=ctk.S)

    _quit = ctk.CTkButton(frame, text="Quit", font=('Segoe Ui', 22), command=exit)
    _quit.pack(padx=10, pady=10)
    _quit.place(relx=0.5, rely=0.52, anchor=ctk.N)

starting()

root.mainloop()
